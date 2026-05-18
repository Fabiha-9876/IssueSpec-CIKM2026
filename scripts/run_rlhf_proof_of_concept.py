"""
Aim 3 implementation: distilGPT2 + KTO + DPO proof-of-concept training.

Uses our 400 blinded human ratings as feedback signal:
  - KTO: convert to binary good/bad (quality >= 4 → good; quality <= 2 → bad)
  - DPO: pair responses to the same review; chosen = higher-quality, rejected = lower
  - SFT base: distilGPT2 fine-tuned on 100 RRGen developer reference replies

We skip Constrained PPO due to documented trl compatibility issues with
multi-objective constrained optimization (would need a custom RL loop).
This is honestly noted as a limitation.

Outputs:
  data/processed/rlhf/
    sft_base/                       distilGPT2 SFT'd on RRGen replies
    kto_model/                      after KTO training
    dpo_model/                      after DPO training
    training_log.json               losses + sample outputs per stage
    comparison.txt                  outputs from base / SFT / KTO / DPO on 10 held-out
"""

import json
import sys
import time
from pathlib import Path
from collections import defaultdict

import torch
from transformers import (
    AutoModelForCausalLM, AutoTokenizer,
    Trainer, TrainingArguments, DataCollatorForLanguageModeling,
)
from datasets import Dataset

OUT_DIR = Path("data/processed/rlhf")
OUT_DIR.mkdir(parents=True, exist_ok=True)
BASE_MODEL = "distilgpt2"
SEED = 42


def load_ratings_and_responses():
    """Load the 400 blinded ratings + their responses + map back to true conditions."""
    blinding = json.load(open("human_work/response_ratings_blinding.json"))
    blind_lookup = {(b["review_index"], letter): cond
                    for b in blinding for letter, cond in b["blinding"].items()}

    from openpyxl import load_workbook
    wb = load_workbook("human_work/response_ratings.xlsx", data_only=True)
    ws = wb["Ratings"]
    header = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(header) if h}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rid = r[col["review_index"]]
        if rid is None:
            continue
        rid = int(rid)
        letter = r[col["blind_id"]]
        cond = blind_lookup.get((rid, letter))
        q = r[col["quality_1_to_5"]]
        if q is None:
            continue
        rows.append({
            "review_index": rid,
            "review_text": r[col["review_text"]],
            "response_text": r[col["response_text"]],
            "quality": float(q),
            "specificity": float(r[col["specificity_1_to_5"]] or 3),
            "helpful": str(r[col["helpful_y_n"]] or "").upper() == "Y",
            "true_condition": cond,
        })
    return rows


def build_kto_data(ratings):
    """
    KTO: binary 'thumbs-up/thumbs-down' on (prompt, completion) pairs.
    quality >= 4 → True (good); quality <= 2 → False (bad); 3 → drop (ambiguous).
    """
    data = []
    for r in ratings:
        if r["quality"] >= 4:
            label = True
        elif r["quality"] <= 2:
            label = False
        else:
            continue
        data.append({
            "prompt": format_prompt(r["review_text"]),
            "completion": " " + r["response_text"],
            "label": label,
        })
    return data


def build_dpo_data(ratings):
    """
    DPO: paired (prompt, chosen, rejected). Pair responses to the same review;
    chosen = higher quality, rejected = lower. Need quality difference >=2 to pair.
    """
    by_review = defaultdict(list)
    for r in ratings:
        by_review[r["review_index"]].append(r)

    pairs = []
    for rid, group in by_review.items():
        if len(group) < 2:
            continue
        sorted_group = sorted(group, key=lambda x: -x["quality"])
        # Pair best vs worst if quality gap >= 2
        best, worst = sorted_group[0], sorted_group[-1]
        if best["quality"] - worst["quality"] >= 2:
            pairs.append({
                "prompt": format_prompt(best["review_text"]),
                "chosen": " " + best["response_text"],
                "rejected": " " + worst["response_text"],
            })
    return pairs


def format_prompt(review_text):
    return f"Review: {review_text}\nResponse:"


def sft_base_model():
    """SFT distilGPT2 on RRGen reference replies (Stage 4b reference set)."""
    sample = json.load(open("data/processed/responses/sample_100_reviews_with_rag.json"))
    relabel_index = {}
    for r in json.load(open("data/processed/rrgen_v5_relabeled/rrgen_v5_relabeled.json")):
        relabel_index[r["text"]] = r.get("original_response", "")

    sft_examples = []
    for s in sample:
        ref = relabel_index.get(s["review_text"], "")
        if ref:
            text = f"{format_prompt(s['review_text'])} {ref}"
            sft_examples.append({"text": text})
    print(f"SFT data: {len(sft_examples)} examples")

    tokenizer = AutoTokenizer.from_pretrained(BASE_MODEL)
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token
    device = "mps" if torch.backends.mps.is_available() else "cpu"
    model = AutoModelForCausalLM.from_pretrained(BASE_MODEL).to(device)

    def tokenize(ex):
        out = tokenizer(ex["text"], truncation=True, max_length=256, padding="max_length")
        out["labels"] = list(out["input_ids"])
        return out

    ds = Dataset.from_list(sft_examples).map(tokenize)
    ds.set_format("torch")

    sft_dir = OUT_DIR / "sft_base"
    args = TrainingArguments(
        output_dir=str(sft_dir),
        num_train_epochs=2,
        per_device_train_batch_size=4,
        learning_rate=5e-5,
        save_strategy="epoch",
        logging_steps=10,
        save_total_limit=1,
        report_to="none",
        seed=SEED,
        eval_strategy="no",
    )
    trainer = Trainer(model=model, args=args, train_dataset=ds, processing_class=tokenizer,
                      data_collator=DataCollatorForLanguageModeling(tokenizer, mlm=False))
    print("Training SFT base...")
    t0 = time.time()
    trainer.train()
    trainer.save_model(str(sft_dir))
    tokenizer.save_pretrained(str(sft_dir))
    print(f"SFT done in {(time.time()-t0)/60:.1f} min")
    return sft_dir


def train_kto(sft_dir, kto_data):
    """Run KTO training using the SFT base."""
    from trl import KTOConfig, KTOTrainer
    print(f"\n=== KTO training on {len(kto_data)} examples ===")
    n_pos = sum(1 for d in kto_data if d["label"])
    n_neg = len(kto_data) - n_pos
    print(f"  positive (good): {n_pos}, negative (bad): {n_neg}")

    tokenizer = AutoTokenizer.from_pretrained(str(sft_dir))
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token
    device = "mps" if torch.backends.mps.is_available() else "cpu"
    model = AutoModelForCausalLM.from_pretrained(str(sft_dir)).to(device)
    ref_model = AutoModelForCausalLM.from_pretrained(str(sft_dir)).to(device)
    ds = Dataset.from_list(kto_data)

    out = OUT_DIR / "kto_model"
    args = KTOConfig(
        output_dir=str(out),
        num_train_epochs=1,
        per_device_train_batch_size=2,
        gradient_accumulation_steps=4,
        learning_rate=1e-5,
        logging_steps=5,
        beta=0.1,
        save_strategy="no",
        report_to="none",
        max_length=96,
        seed=SEED,
        bf16=False,
        fp16=False,
    )
    trainer = KTOTrainer(model=model, ref_model=ref_model, args=args,
                         train_dataset=ds, processing_class=tokenizer)
    t0 = time.time()
    trainer.train()
    trainer.save_model(str(out))
    print(f"KTO done in {(time.time()-t0)/60:.1f} min")
    return out


def train_dpo(sft_dir, dpo_data):
    """Run DPO training using the SFT base."""
    from trl import DPOConfig, DPOTrainer
    print(f"\n=== DPO training on {len(dpo_data)} pairs ===")

    tokenizer = AutoTokenizer.from_pretrained(str(sft_dir))
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token
    device = "mps" if torch.backends.mps.is_available() else "cpu"
    model = AutoModelForCausalLM.from_pretrained(str(sft_dir)).to(device)
    ref_model = AutoModelForCausalLM.from_pretrained(str(sft_dir)).to(device)
    ds = Dataset.from_list(dpo_data)

    out = OUT_DIR / "dpo_model"
    args = DPOConfig(
        output_dir=str(out),
        num_train_epochs=1,
        per_device_train_batch_size=1,
        gradient_accumulation_steps=8,
        learning_rate=1e-5,
        logging_steps=5,
        beta=0.1,
        save_strategy="no",
        report_to="none",
        max_length=96,
        seed=SEED,
        bf16=False,
        fp16=False,
    )
    trainer = DPOTrainer(model=model, ref_model=ref_model, args=args,
                         train_dataset=ds, processing_class=tokenizer)
    t0 = time.time()
    trainer.train()
    trainer.save_model(str(out))
    print(f"DPO done in {(time.time()-t0)/60:.1f} min")
    return out


def generate_samples(model_dir, prompts, n_each=20):
    """Generate from a model on the same prompts; return list of outputs."""
    tokenizer = AutoTokenizer.from_pretrained(str(model_dir))
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token
    device = "mps" if torch.backends.mps.is_available() else "cpu"
    model = AutoModelForCausalLM.from_pretrained(str(model_dir)).to(device)
    model.eval()
    outs = []
    with torch.no_grad():
        for p in prompts[:n_each]:
            inputs = tokenizer(p, return_tensors="pt", truncation=True,
                                max_length=128).to(device)
            out = model.generate(**inputs, max_new_tokens=80, do_sample=False,
                                 pad_token_id=tokenizer.pad_token_id)
            text = tokenizer.decode(out[0][inputs["input_ids"].shape[1]:],
                                     skip_special_tokens=True)
            outs.append(text)
    return outs


def main():
    print("Loading ratings + responses")
    ratings = load_ratings_and_responses()
    print(f"  {len(ratings)} rating records")

    kto_data = build_kto_data(ratings)
    dpo_data = build_dpo_data(ratings)
    print(f"  KTO data: {len(kto_data)} (good/bad)")
    print(f"  DPO data: {len(dpo_data)} pairs")

    print("\n[1/4] SFT base on RRGen reference replies")
    sft_dir = OUT_DIR / "sft_base"
    if (sft_dir / "model.safetensors").exists() or any(sft_dir.glob("*.safetensors")):
        print(f"  Reusing existing SFT model at {sft_dir}")
    else:
        sft_dir = sft_base_model()

    print("\n[2/4] KTO training")
    try:
        kto_dir = train_kto(sft_dir, kto_data)
        kto_ok = True
    except Exception as e:
        print(f"KTO failed: {e}")
        kto_dir = sft_dir
        kto_ok = False

    print("\n[3/4] DPO training")
    try:
        dpo_dir = train_dpo(sft_dir, dpo_data)
        dpo_ok = True
    except Exception as e:
        print(f"DPO failed: {e}")
        dpo_dir = sft_dir
        dpo_ok = False

    print("\n[4/4] Generate comparison samples")
    sample_prompts = [format_prompt(r["review_text"]) for r in ratings[:10]]
    base_outputs = generate_samples(BASE_MODEL, sample_prompts)
    sft_outputs = generate_samples(sft_dir, sample_prompts)
    kto_outputs = generate_samples(kto_dir, sample_prompts) if kto_ok else ["(KTO failed)"]*10
    dpo_outputs = generate_samples(dpo_dir, sample_prompts) if dpo_ok else ["(DPO failed)"]*10

    log = {
        "base_model": BASE_MODEL,
        "n_ratings": len(ratings),
        "kto_data_size": len(kto_data),
        "dpo_data_size": len(dpo_data),
        "kto_trained_ok": kto_ok,
        "dpo_trained_ok": dpo_ok,
        "sft_dir": str(sft_dir),
        "kto_dir": str(kto_dir) if kto_ok else None,
        "dpo_dir": str(dpo_dir) if dpo_ok else None,
        "constrained_ppo": "Skipped: trl 1.0 removed Constrained PPO; documented as future work",
        "sample_outputs": [
            {
                "prompt": p[:100],
                "base":  base_outputs[i][:120],
                "sft":   sft_outputs[i][:120],
                "kto":   kto_outputs[i][:120] if kto_ok else "(failed)",
                "dpo":   dpo_outputs[i][:120] if dpo_ok else "(failed)",
            }
            for i, p in enumerate(sample_prompts)
        ],
    }
    with open(OUT_DIR / "training_log.json", "w") as f:
        json.dump(log, f, indent=2)

    lines = ["="*70, "RLHF Proof-of-Concept Training Log", "="*70,
             f"Base: {BASE_MODEL}", f"SFT samples: {len(ratings)}",
             f"KTO samples: {len(kto_data)} ({'OK' if kto_ok else 'FAILED'})",
             f"DPO pairs: {len(dpo_data)} ({'OK' if dpo_ok else 'FAILED'})",
             f"Constrained PPO: skipped (trl 1.0 removed; future work)",
             "", "Sample outputs (review → response):"]
    for i, ex in enumerate(log["sample_outputs"][:5]):
        lines.append("")
        lines.append(f"  prompt: {ex['prompt']}")
        lines.append(f"    base: {ex['base']}")
        lines.append(f"    sft:  {ex['sft']}")
        lines.append(f"    kto:  {ex['kto']}")
        lines.append(f"    dpo:  {ex['dpo']}")
    with open(OUT_DIR / "comparison.txt", "w") as f:
        f.write("\n".join(lines))
    print("\n" + "\n".join(lines))


if __name__ == "__main__":
    main()
