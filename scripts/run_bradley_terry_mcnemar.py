"""
Bradley-Terry strengths + McNemar pairwise tests on Exp 2 paired data.

Proposal §8 specifies these tests for Exp 3 (RLHF variants), but end-to-end
RLHF training with human eval was deferred (§5.5). We therefore apply the
same statistical machinery to the Stage 4b 4-condition human evaluation,
where we *do* have row-level paired ratings on the same 100 reviews.

Bradley-Terry: per-condition strength theta from pairwise quality wins.
McNemar: on the helpful_y_n binary outcome, paired across conditions.

Output: data/processed/experiments/exp3_bt_mcnemar.json
"""

import json
from collections import defaultdict
from itertools import combinations
from pathlib import Path

import numpy as np
import choix
from openpyxl import load_workbook
from statsmodels.stats.contingency_tables import mcnemar

REPORTING_ORDER = ["rrgen_baseline", "prompt_baseline", "reviewagent_no_spec", "reviewagent_full"]


def load_paired():
    blinding = json.load(open("human_work/response_ratings_blinding.json"))
    blind_lookup = {}
    for entry in blinding:
        for letter, cond in entry["blinding"].items():
            blind_lookup[(entry["review_index"], letter)] = cond

    wb = load_workbook("human_work/response_ratings.xlsx", data_only=True)
    ws = wb["Ratings"]
    header = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(header) if h}

    paired = defaultdict(dict)
    for r in ws.iter_rows(min_row=2, values_only=True):
        rid = r[col["review_index"]]
        if rid is None:
            continue
        rid = int(rid)
        true_cond = blind_lookup.get((rid, r[col["blind_id"]]))
        if true_cond is None:
            continue
        q = r[col["quality_1_to_5"]]
        h = str(r[col["helpful_y_n"]] or "").strip().upper()
        if q is None:
            continue
        try:
            q = float(q)
        except Exception:
            continue
        paired[rid][true_cond] = {"q": q, "h": h}
    return paired


def bradley_terry(paired):
    """Build pairwise wins from quality scores then fit BT via choix MLE."""
    cond_idx = {c: i for i, c in enumerate(REPORTING_ORDER)}
    pairs_won = []
    ties = 0
    for rid, conds in paired.items():
        for a, b in combinations(REPORTING_ORDER, 2):
            if a not in conds or b not in conds:
                continue
            qa, qb = conds[a]["q"], conds[b]["q"]
            if qa > qb:
                pairs_won.append((cond_idx[a], cond_idx[b]))
            elif qb > qa:
                pairs_won.append((cond_idx[b], cond_idx[a]))
            else:
                ties += 1
    n = len(REPORTING_ORDER)
    # MLE with mild regularization for stability
    params = choix.ilsr_pairwise(n, pairs_won, alpha=0.01)
    return params, len(pairs_won), ties


def mcnemar_helpfulness(paired):
    """Paired McNemar on helpful_y_n for every condition pair."""
    out = {}
    for a, b in combinations(REPORTING_ORDER, 2):
        a_only = b_only = both = neither = 0
        for rid, conds in paired.items():
            if a not in conds or b not in conds:
                continue
            ha = conds[a]["h"] == "Y"
            hb = conds[b]["h"] == "Y"
            if ha and hb: both += 1
            elif ha and not hb: a_only += 1
            elif hb and not ha: b_only += 1
            else: neither += 1
        # 2x2 table:  rows = a (Y, N), cols = b (Y, N)
        table = [[both, a_only], [b_only, neither]]
        try:
            res = mcnemar(table, exact=False, correction=True)
            stat, p = float(res.statistic), float(res.pvalue)
        except Exception:
            stat, p = float("nan"), float("nan")
        out[f"{a}_vs_{b}"] = {
            "table": table,
            "a_helpful_only": a_only,
            "b_helpful_only": b_only,
            "both_helpful": both,
            "neither_helpful": neither,
            "mcnemar_chi2": stat,
            "p_value": p,
        }
    return out


def main():
    paired = load_paired()
    print(f"Loaded {len(paired)} paired reviews across {len(REPORTING_ORDER)} conditions.")

    # Bradley-Terry
    params, n_wins, n_ties = bradley_terry(paired)
    print(f"\nBradley-Terry MLE on {n_wins} pairwise wins ({n_ties} ties dropped):")
    rank_idx = np.argsort(-params)
    for r, idx in enumerate(rank_idx, 1):
        print(f"  rank {r}: {REPORTING_ORDER[idx]:25s} theta = {params[idx]:+.3f}")

    # McNemar on helpfulness
    mcn = mcnemar_helpfulness(paired)
    print(f"\nMcNemar on helpful_y_n (chi-squared, continuity-corrected):")
    print(f"  {'pair':50s} {'a-only':>8} {'b-only':>8} {'chi2':>8} {'p':>10}")
    for pair, r in mcn.items():
        sig = "***" if r["p_value"] < 0.001 else ("**" if r["p_value"] < 0.01 else ("*" if r["p_value"] < 0.05 else ""))
        print(f"  {pair:50s} {r['a_helpful_only']:>8} {r['b_helpful_only']:>8} "
              f"{r['mcnemar_chi2']:>8.3f} {r['p_value']:>10.3e} {sig}")

    out = {
        "applied_to": "Stage 4b 4-condition human eval (Exp 3 RLHF training was deferred per §5.5)",
        "n_paired": len(paired),
        "conditions": REPORTING_ORDER,
        "bradley_terry": {
            "method": "ILSR MLE (choix.ilsr_pairwise, alpha=0.01)",
            "n_pairwise_wins": n_wins,
            "n_ties_dropped": n_ties,
            "theta": dict(zip(REPORTING_ORDER, [round(float(x), 4) for x in params])),
            "ranking": [REPORTING_ORDER[i] for i in rank_idx],
        },
        "mcnemar_helpfulness": {k: {kk: (vv if not isinstance(vv, list) else vv) for kk, vv in v.items()} for k, v in mcn.items()},
    }
    Path("data/processed/experiments").mkdir(parents=True, exist_ok=True)
    out_path = Path("data/processed/experiments/exp3_bt_mcnemar.json")
    with open(out_path, "w") as f:
        json.dump(out, f, indent=2)
    print(f"\nSaved {out_path}")


if __name__ == "__main__":
    main()
