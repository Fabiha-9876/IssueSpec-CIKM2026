"""
Friedman + post-hoc Nemenyi test on the Exp 2 quality ratings.

Proposal §8 specifies these tests across the 4 conditions
(rrgen_baseline, prompt_baseline, reviewagent_no_spec, reviewagent_full).

Loads the same blinded human-work spreadsheet that score_human_work.py reads,
builds the 100x4 paired matrix of quality scores, and reports:
  - Friedman omnibus statistic + p-value
  - Pairwise Nemenyi p-values (k=4, N=n_paired)
  - Critical difference (CD) at alpha=0.05 for visual ranking

Output: data/processed/experiments/exp2_friedman_nemenyi.json
"""

import json
from collections import defaultdict
from pathlib import Path

import numpy as np
import scipy.stats as st
import scikit_posthocs as sp
from openpyxl import load_workbook

REPORTING_ORDER = ["rrgen_baseline", "prompt_baseline", "reviewagent_no_spec", "reviewagent_full"]


def load_paired_quality():
    blinding = json.load(open("human_work/response_ratings_blinding.json"))
    blind_lookup = {}
    for entry in blinding:
        for letter, cond in entry["blinding"].items():
            blind_lookup[(entry["review_index"], letter)] = cond

    wb = load_workbook("human_work/response_ratings.xlsx", data_only=True)
    ws = wb["Ratings"]
    header = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(header) if h}

    paired = defaultdict(dict)
    for r in ws.iter_rows(min_row=2, values_only=True):
        rid = r[col["review_index"]]
        if rid is None:
            continue
        rid = int(rid)
        true_cond = blind_lookup.get((rid, r[col["blind_id"]]))
        if true_cond is None:
            continue
        q = r[col["quality_1_to_5"]]
        if q is None:
            continue
        try:
            q = float(q)
        except Exception:
            continue
        paired[rid][true_cond] = q

    rows = []
    for rid, conds in paired.items():
        if all(c in conds for c in REPORTING_ORDER):
            rows.append([conds[c] for c in REPORTING_ORDER])
    return np.array(rows)


def main():
    matrix = load_paired_quality()
    print(f"Paired matrix shape: {matrix.shape}  (rows=reviews, cols=4 conditions)")
    print(f"Conditions in column order: {REPORTING_ORDER}")
    print(f"Mean quality per condition:")
    for c, m in zip(REPORTING_ORDER, matrix.mean(axis=0)):
        print(f"  {c:25s} {m:.3f}")

    # Friedman omnibus
    chi2, p = st.friedmanchisquare(*[matrix[:, i] for i in range(matrix.shape[1])])
    print(f"\nFriedman: chi2 = {chi2:.3f}, df = {matrix.shape[1]-1}, p = {p:.3e}")

    # Post-hoc Nemenyi (Wilcoxon-style ranks)
    nemenyi = sp.posthoc_nemenyi_friedman(matrix)
    nemenyi.index = REPORTING_ORDER
    nemenyi.columns = REPORTING_ORDER
    print("\nNemenyi pairwise p-values:")
    print(nemenyi.round(4).to_string())

    # Critical difference at alpha=0.05 (Demsar 2006 formulation)
    k = matrix.shape[1]
    n = matrix.shape[0]
    # q_alpha for k=4 at alpha=0.05 (from studentized range distribution table) is 2.569
    q_alpha = 2.569 if k == 4 else None
    if q_alpha is not None:
        cd = q_alpha * np.sqrt(k * (k + 1) / (6.0 * n))
        print(f"\nCritical difference at alpha=0.05 (k={k}, N={n}): {cd:.4f}")
    else:
        cd = None

    # Mean rank (lower is better in Demsar convention -> we negate so higher quality = lower rank)
    # Compute average ranks for the *negated* matrix so the best condition has rank 1.
    ranks = st.rankdata(-matrix, axis=1)
    mean_ranks = ranks.mean(axis=0)
    print(f"\nMean ranks (1 = best, {k} = worst):")
    for c, r in zip(REPORTING_ORDER, mean_ranks):
        print(f"  {c:25s} {r:.3f}")

    out = {
        "n_paired": int(n),
        "k_conditions": int(k),
        "conditions": REPORTING_ORDER,
        "mean_quality": dict(zip(REPORTING_ORDER, [round(float(x), 3) for x in matrix.mean(axis=0)])),
        "friedman": {"chi2": float(chi2), "df": k - 1, "p_value": float(p)},
        "nemenyi": {a: {b: float(nemenyi.loc[a, b]) for b in REPORTING_ORDER} for a in REPORTING_ORDER},
        "critical_difference_alpha_0.05": float(cd) if cd is not None else None,
        "mean_ranks": dict(zip(REPORTING_ORDER, [round(float(x), 3) for x in mean_ranks])),
    }
    Path("data/processed/experiments").mkdir(parents=True, exist_ok=True)
    out_path = Path("data/processed/experiments/exp2_friedman_nemenyi.json")
    with open(out_path, "w") as f:
        json.dump(out, f, indent=2)
    print(f"\nSaved {out_path}")


if __name__ == "__main__":
    main()
