"""
Compute paired effect sizes and bootstrap CIs for the RQ2 +2.36 result (lever #2).

Reads the 400-row lead-author response_ratings.xlsx, pairs full vs no_spec on
the same 100 reviews, and reports:

  - Mean delta with paired-bootstrap 95% CI (B=10,000)
  - Cohen's d_z (paired)
  - Cohen's d_s (independent-sample, for cross-paper comparability)
  - Cliff's delta (independent-sample, non-parametric effect size)
  - Matched-pair sign breakdown (n+/n-/n=)
  - For reference: paired Wilcoxon p-value (from scipy)

The script also computes the same effect sizes on the three other paired
contrasts (full vs rrgen, full vs prompt, no_spec vs prompt) for completeness.

Run:  python3 scripts/compute_effect_sizes_rq2.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

try:
    from scipy.stats import wilcoxon
except ImportError:
    print("pip install scipy", file=sys.stderr); sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
RATINGS_XLSX = ROOT / "human_work/response_ratings.xlsx"
BLINDING_JSON = ROOT / "human_work/response_ratings_blinding.json"
OUT_PATH = ROOT / "data/processed/responses/effect_sizes_rq2.json"

CONDS = ("rrgen_baseline", "prompt_baseline", "reviewagent_no_spec", "reviewagent_full")
B_BOOT = 10000
SEED = 2026


def load_ratings() -> dict[str, dict[int, int]]:
    """Return condition -> {review_index -> quality_score}."""
    wb = load_workbook(RATINGS_XLSX, data_only=True)
    ws = wb["Ratings"]
    blinding = {b["review_index"]: b["blinding"]
                for b in json.load(open(BLINDING_JSON))}

    out: dict[str, dict[int, int]] = {c: {} for c in CONDS}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ri, _cid, _it, _rt, bid, _resp, q, _sp, _h, _nt = row
        if q is None:
            continue
        cond = blinding[ri][bid]
        out[cond][ri] = int(q)
    return out


def paired_vectors(by_cond: dict[str, dict[int, int]], a: str, b: str
                    ) -> tuple[np.ndarray, np.ndarray]:
    """Aligned paired vectors (a, b) over the intersection of review indices."""
    common = sorted(set(by_cond[a]) & set(by_cond[b]))
    return (np.array([by_cond[a][i] for i in common], dtype=float),
            np.array([by_cond[b][i] for i in common], dtype=float))


def cliffs_delta(x: np.ndarray, y: np.ndarray) -> float:
    """Independent-sample Cliff's delta: (#x>y - #x<y) / (|x| |y|)."""
    diff = x[:, None] - y[None, :]
    n_pos = int(np.sum(diff > 0))
    n_neg = int(np.sum(diff < 0))
    return (n_pos - n_neg) / (len(x) * len(y))


def cohens_d_paired(diff: np.ndarray) -> float:
    """d_z: mean(diff) / sd(diff) using ddof=1."""
    sd = float(np.std(diff, ddof=1))
    if sd == 0:
        return float("inf") if float(np.mean(diff)) != 0 else 0.0
    return float(np.mean(diff)) / sd


def cohens_d_independent(x: np.ndarray, y: np.ndarray) -> float:
    """Independent-sample Cohen's d_s using pooled SD."""
    nx, ny = len(x), len(y)
    sx2, sy2 = float(np.var(x, ddof=1)), float(np.var(y, ddof=1))
    pooled = np.sqrt(((nx - 1) * sx2 + (ny - 1) * sy2) / (nx + ny - 2))
    if pooled == 0:
        return 0.0
    return (float(np.mean(x)) - float(np.mean(y))) / pooled


def paired_bootstrap_ci(diff: np.ndarray, b: int, alpha: float,
                          rng: np.random.Generator
                          ) -> tuple[float, float]:
    n = len(diff)
    idx = rng.integers(0, n, size=(b, n))
    means = diff[idx].mean(axis=1)
    lo = float(np.quantile(means, alpha / 2))
    hi = float(np.quantile(means, 1 - alpha / 2))
    return lo, hi


def cliffs_label(delta: float) -> str:
    """Romano et al. (2006) magnitude bands."""
    a = abs(delta)
    if a < 0.147: return "negligible"
    if a < 0.33:  return "small"
    if a < 0.474: return "medium"
    return "large"


def analyze_pair(by_cond: dict, treat: str, control: str,
                  rng: np.random.Generator) -> dict:
    x, y = paired_vectors(by_cond, treat, control)
    diff = x - y
    n = len(diff)
    n_pos = int(np.sum(diff > 0))
    n_neg = int(np.sum(diff < 0))
    n_eq = int(np.sum(diff == 0))
    d_z = cohens_d_paired(diff)
    d_s = cohens_d_independent(x, y)
    cd = cliffs_delta(x, y)
    lo, hi = paired_bootstrap_ci(diff, b=B_BOOT, alpha=0.05, rng=rng)
    try:
        wp = float(wilcoxon(x, y, zero_method="pratt").pvalue)
    except ValueError:
        wp = None
    return {
        "treatment": treat, "control": control, "n_pairs": n,
        "mean_treatment": float(np.mean(x)),
        "mean_control": float(np.mean(y)),
        "mean_diff": float(np.mean(diff)),
        "sd_diff": float(np.std(diff, ddof=1)),
        "ci95_paired_bootstrap": [lo, hi],
        "cohens_d_paired": d_z,
        "cohens_d_independent": d_s,
        "cliffs_delta": cd,
        "cliffs_magnitude": cliffs_label(cd),
        "matched_sign_breakdown": {"treat_higher": n_pos,
                                     "control_higher": n_neg,
                                     "tie": n_eq},
        "wilcoxon_p_two_sided": wp,
        "bootstrap_iterations": B_BOOT,
    }


def main() -> int:
    by_cond = load_ratings()
    rng = np.random.default_rng(SEED)

    contrasts = [
        ("reviewagent_full", "reviewagent_no_spec"),  # headline +2.36
        ("reviewagent_full", "prompt_baseline"),
        ("reviewagent_full", "rrgen_baseline"),
        ("reviewagent_no_spec", "prompt_baseline"),  # negative-finding contrast
    ]
    results = {"headline_pair": "reviewagent_full vs reviewagent_no_spec",
               "n_reviews": 100, "rubric_scale": "1-5",
               "bootstrap_seed": SEED,
               "contrasts": [analyze_pair(by_cond, t, c, rng) for t, c in contrasts]}

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w") as f:
        json.dump(results, f, indent=2)
    print(json.dumps(results, indent=2))
    print(f"\nsaved -> {OUT_PATH}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
