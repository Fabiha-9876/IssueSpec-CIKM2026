"""
Prepare Full RRGen Dataset for Annotator Review
=================================================

Creates:
1. Full CSV (all 215K) - for reference and bulk import
2. Annotation Excel sheets - stratified samples for human review
   - Sample per label category (balanced)
   - All low-confidence reviews (need most attention)
   - Random sample of high-confidence (spot-check)

Usage:
    python3 scripts/prepare_annotation_dataset.py

Output:
    <HOME>/Desktop/Review Agent/
        RRGen_Full_Annotator_Review.xlsx     - Excel for annotators
        RRGen_Full_Dataset.csv               - Full 215K CSV
"""

import json
import csv
import sys
import random
import argparse
from pathlib import Path
from collections import Counter

import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.stage1.classifier import LABELS

# Styles
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
VERIFY_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def create_instructions_sheet(wb, total, sample_size, stats):
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = "2F5496"
    ws.column_dimensions["A"].width = 100

    lines = [
        ("ANNOTATOR REVIEW INSTRUCTIONS", True),
        ("", False),
        (f"DATASET: Full RRGen Dataset ({total:,} reviews, {sample_size:,} sampled for review)", False),
        ("", False),
        ("WHAT YOU NEED TO DO:", True),
        ("1. Go to each category tab (bug_report, feature_request, etc.)", False),
        ("2. For each review, check if the predicted label is correct", False),
        ("3. Fill the yellow 'Correct? (Y/N)' column", False),
        ("4. If N, select the correct label in 'Correct Label' column", False),
        ("5. Optionally add comments for ambiguous cases", False),
        ("", False),
        ("CATEGORY DEFINITIONS:", True),
        ("  bug_report      - Crashes, errors, broken features, malfunctions", False),
        ("  feature_request  - Requests for new features or improvements", False),
        ("  performance      - Speed, battery, memory, lag, loading times", False),
        ("  usability        - Confusing UI, hard to use, poor navigation", False),
        ("  compatibility    - Device-specific or OS-specific issues", False),
        ("  praise           - Positive feedback, compliments", False),
        ("  other            - Doesn't fit any category above", False),
        ("", False),
        ("TIPS:", True),
        ("  - 'slow'/'lag' = performance (NOT bug_report)", False),
        ("  - 'crash on my Samsung' = compatibility (device-specific)", False),
        ("  - 'crash' (no device) = bug_report", False),
        ("  - Short praise like 'good app' = praise", False),
        ("  - If review could fit 2 categories, check if predicted one is reasonable", False),
        ("", False),
        ("DATASET STATISTICS:", True),
    ]

    for label in LABELS:
        count = stats.get(label, 0)
        pct = count / total * 100
        lines.append((f"  {label:20s}: {count:6,} ({pct:5.1f}%)", False))

    lines.append(("", False))
    lines.append((f"  Total:              {total:,}", True))
    lines.append((f"  Sampled for review: {sample_size:,}", True))
    lines.append(("", False))
    lines.append(("SHEETS IN THIS WORKBOOK:", True))
    lines.append(("  - Instructions (this sheet)", False))
    lines.append(("  - One tab per category with sampled reviews", False))
    lines.append(("  - Summary tab with overall statistics", False))

    for i, (text, bold) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=12 if bold else 11)
        cell.alignment = WRAP_ALIGN


def create_category_sheet(wb, label, reviews, yn_validation, label_validation):
    ws = wb.create_sheet(label)

    # Color code by category
    colors = {
        "bug_report": "FF6B6B", "feature_request": "4ECDC4",
        "performance": "FFE66D", "usability": "A8E6CF",
        "compatibility": "DDA0DD", "praise": "87CEEB", "other": "D3D3D3",
    }
    ws.sheet_properties.tabColor = colors.get(label, "808080")

    headers = [
        "ID", "Review Text", "Rating", "Predicted Label",
        "Confidence", "Needs HITL",
        "Correct? (Y/N)", "Correct Label (if N)", "Comments"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    widths = [8, 70, 8, 16, 12, 10, 14, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, r in enumerate(reviews, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=r["text"]).alignment = WRAP_ALIGN
        ws.cell(row=i, column=3, value=r.get("rating", "")).alignment = CENTER_ALIGN
        ws.cell(row=i, column=4, value=r.get("predicted_label", "")).alignment = CENTER_ALIGN
        ws.cell(row=i, column=5, value=r.get("confidence", 0)).alignment = CENTER_ALIGN
        ws.cell(row=i, column=6, value="Yes" if r.get("needs_hitl") else "No").alignment = CENTER_ALIGN

        yn_cell = ws.cell(row=i, column=7)
        yn_cell.fill = VERIFY_FILL
        yn_cell.alignment = CENTER_ALIGN
        yn_validation.add(yn_cell)

        label_cell = ws.cell(row=i, column=8)
        label_cell.fill = VERIFY_FILL
        label_cell.alignment = CENTER_ALIGN
        label_validation.add(label_cell)

        comment_cell = ws.cell(row=i, column=9)
        comment_cell.fill = VERIFY_FILL
        comment_cell.alignment = WRAP_ALIGN

        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = THIN_BORDER

    ws.freeze_panes = "A2"
    return ws


def create_summary_sheet(wb, all_reviews, sampled_counts):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "FFC000"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    ws.cell(row=1, column=1, value="Dataset Summary").font = Font(bold=True, size=14)

    headers = ["Label", "Total Count", "Sampled", "Sample %"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 4)

    total_counts = Counter(r["predicted_label"] for r in all_reviews)

    for i, label in enumerate(LABELS, 4):
        total = total_counts.get(label, 0)
        sampled = sampled_counts.get(label, 0)
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=total)
        ws.cell(row=i, column=3, value=sampled)
        ws.cell(row=i, column=4, value=f"{sampled/total*100:.1f}%" if total > 0 else "N/A")

    row = 4 + len(LABELS) + 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(all_reviews)).font = Font(bold=True)
    ws.cell(row=row, column=3, value=sum(sampled_counts.values())).font = Font(bold=True)

    # Confidence breakdown
    row += 2
    ws.cell(row=row, column=1, value="Confidence Breakdown").font = Font(bold=True, size=14)
    row += 2
    confs = [r["confidence"] for r in all_reviews]
    brackets = [
        ("High (>=0.9)", sum(1 for c in confs if c >= 0.9)),
        ("Good (0.8-0.9)", sum(1 for c in confs if 0.8 <= c < 0.9)),
        ("Medium (0.6-0.8)", sum(1 for c in confs if 0.6 <= c < 0.8)),
        ("Low (0.4-0.6)", sum(1 for c in confs if 0.4 <= c < 0.6)),
        ("Very Low (<0.4)", sum(1 for c in confs if c < 0.4)),
    ]
    for label, count in brackets:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{count/len(confs)*100:.1f}%")
        row += 1


def main():
    parser = argparse.ArgumentParser(description="Prepare annotation dataset")
    parser.add_argument("--input", default="data/processed/rrgen_full_labeled/rrgen_full_labeled.json")
    parser.add_argument("--samples-per-label", type=int, default=500,
                        help="Max reviews to sample per label for annotation (default: 500)")
    parser.add_argument("--low-conf-all", action="store_true", default=True,
                        help="Include ALL low-confidence reviews (default: True)")
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    random.seed(args.seed)
    np.random.seed(args.seed)

    output_dir = Path("<HOME>/Desktop/Review Agent")

    # Load data
    print("Loading labeled data...")
    with open(args.input) as f:
        all_reviews = json.load(f)
    print(f"  Total reviews: {len(all_reviews):,}")

    # ── Step 1: Export full CSV ──────────────────────────────────────────
    print("\nStep 1: Exporting full CSV...")
    csv_path = output_dir / "RRGen_Full_Dataset.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "id", "text", "rating", "predicted_label", "confidence",
            "needs_hitl", "app_id", "timestamp",
            "bug_report_conf", "feature_request_conf", "performance_conf",
            "usability_conf", "compatibility_conf", "praise_conf", "other_conf",
        ])
        for i, r in enumerate(all_reviews, 1):
            confs = r.get("all_confidences", {})
            writer.writerow([
                i, r["text"], r.get("rating", ""), r["predicted_label"],
                r["confidence"], r.get("needs_hitl", False),
                r.get("app_id", ""), r.get("timestamp", ""),
                *[confs.get(l, 0) for l in LABELS],
            ])
    print(f"  Saved: {csv_path} ({len(all_reviews):,} rows)")

    # ── Step 2: Stratified sampling for annotation ───────────────────────
    print("\nStep 2: Stratified sampling for annotation...")

    # Group by label
    by_label = {}
    for r in all_reviews:
        label = r["predicted_label"]
        if label not in by_label:
            by_label[label] = []
        by_label[label].append(r)

    # Sample strategy:
    # - For each label: take up to N samples, prioritizing low-confidence
    # - For tail classes (performance, compatibility): take ALL of them
    sampled = {}
    sampled_counts = {}

    for label in LABELS:
        reviews = by_label.get(label, [])
        if not reviews:
            sampled[label] = []
            sampled_counts[label] = 0
            continue

        # Sort by confidence (low first — most uncertain, need review most)
        reviews_sorted = sorted(reviews, key=lambda r: r["confidence"])

        # Tail classes: take ALL
        if label in ("performance", "compatibility"):
            sample = reviews_sorted  # all of them
        elif len(reviews_sorted) <= args.samples_per_label:
            sample = reviews_sorted  # fewer than limit, take all
        else:
            # Take lowest confidence ones first (most need review)
            # 60% low confidence + 40% random from rest
            low_conf_count = int(args.samples_per_label * 0.6)
            random_count = args.samples_per_label - low_conf_count

            low_conf = reviews_sorted[:low_conf_count]
            remaining = reviews_sorted[low_conf_count:]
            random_sample = random.sample(remaining, min(random_count, len(remaining)))
            sample = low_conf + random_sample

        sampled[label] = sample
        sampled_counts[label] = len(sample)
        print(f"    {label:20s}: {len(sample):5d} sampled (of {len(reviews):,})")

    total_sampled = sum(sampled_counts.values())
    print(f"    {'TOTAL':20s}: {total_sampled:5d} sampled for annotation")

    # ── Step 3: Create Excel workbook ────────────────────────────────────
    print(f"\nStep 3: Creating annotation Excel...")
    wb = Workbook()

    label_counts = Counter(r["predicted_label"] for r in all_reviews)
    create_instructions_sheet(wb, len(all_reviews), total_sampled, label_counts)

    # Data validations (shared across sheets)
    yn_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    yn_validation.error = "Please enter Y or N"
    label_validation = DataValidation(
        type="list", formula1='"' + ",".join(LABELS) + '"', allow_blank=True
    )
    label_validation.error = "Please select a valid label"

    for label in LABELS:
        reviews = sampled.get(label, [])
        if reviews:
            # Need new validation objects per sheet
            yn_val = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
            label_val = DataValidation(
                type="list", formula1='"' + ",".join(LABELS) + '"', allow_blank=True
            )
            ws = create_category_sheet(wb, label, reviews, yn_val, label_val)
            ws.add_data_validation(yn_val)
            ws.add_data_validation(label_val)
            print(f"    Created sheet: {label} ({len(reviews)} reviews)")

    create_summary_sheet(wb, all_reviews, sampled_counts)

    xlsx_path = output_dir / "RRGen_Full_Annotator_Review.xlsx"
    print(f"\n  Saving Excel...")
    wb.save(xlsx_path)
    print(f"  Saved: {xlsx_path}")

    # ── Summary ──────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"DONE! Annotation dataset ready.")
    print(f"{'='*60}")
    print(f"\n  Files created:")
    print(f"    1. {csv_path}")
    print(f"       Full dataset: {len(all_reviews):,} reviews")
    print(f"    2. {xlsx_path}")
    print(f"       Annotation sample: {total_sampled:,} reviews")
    print(f"       (one tab per category, sorted by confidence)")
    print(f"\n  For annotators: Open the Excel file, fill Y/N in yellow columns")
    print(f"  For training: Use the CSV for the full dataset")


if __name__ == "__main__":
    main()
