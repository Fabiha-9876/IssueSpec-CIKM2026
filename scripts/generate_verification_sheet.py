"""
Generate Human Verification Spreadsheet from LLM-Labeled Reviews
================================================================

Takes LLM-labeled RRGen reviews and exports an Excel spreadsheet where
human reviewers just verify: "Is this label correct? Y/N"

If N, they provide the correct label.

This is MUCH faster than manual labeling from scratch:
  - Manual labeling: 30-60 sec/review (read + think + decide)
  - Verification: 5-10 sec/review (read + confirm/reject LLM's answer)

Usage:
    python3 scripts/generate_verification_sheet.py
    python3 scripts/generate_verification_sheet.py --output my_sheet.xlsx
    python3 scripts/generate_verification_sheet.py --samples 1000  # first 1000 only

Output:
    Synthetic_Data_Verification_RRGen.xlsx (or custom name)
"""

import json
import sys
import argparse
from pathlib import Path
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.stage1.classifier import LABELS


# ══════════════════════════════════════════════════════════════════════════════
# Styles
# ══════════════════════════════════════════════════════════════════════════════

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
VERIFY_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # yellow
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: Instructions
# ══════════════════════════════════════════════════════════════════════════════

def create_instructions_sheet(wb):
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = "2F5496"
    ws.column_dimensions["A"].width = 100

    instructions = [
        ("VERIFICATION INSTRUCTIONS", True),
        ("", False),
        ("PURPOSE:", True),
        ("An LLM (GPT-4o-mini) has pre-labeled app store reviews into categories.", False),
        ("Your job is to VERIFY whether each label is correct — NOT to label from scratch.", False),
        ("This should take 5-10 seconds per review.", False),
        ("", False),
        ("HOW TO VERIFY:", True),
        ("1. Read the review text", False),
        ("2. Read the LLM's predicted label and its reasoning", False),
        ("3. In the 'Correct? (Y/N)' column:", False),
        ("   - Type Y if the label is correct", False),
        ("   - Type N if the label is wrong", False),
        ("4. If N, select the correct label in the 'Correct Label' column", False),
        ("5. Optionally add a comment if the review is ambiguous", False),
        ("", False),
        ("CATEGORY DEFINITIONS:", True),
        ("bug_report      — Crashes, errors, broken features, malfunctions", False),
        ("feature_request  — Requests for new features or improvements", False),
        ("performance      — Speed, battery, memory, lag, loading times", False),
        ("usability        — Confusing UI, hard to use, poor navigation", False),
        ("compatibility    — Device-specific or OS-specific issues", False),
        ("praise           — Positive feedback, compliments", False),
        ("other            — Doesn't fit any above category", False),
        ("", False),
        ("TIPS:", True),
        ("- If a review could fit multiple categories, check if the LLM picked a reasonable one", False),
        ("- 'slow'/'lag' = performance, NOT bug_report", False),
        ("- 'crash on my Samsung' = compatibility (device-specific)", False),
        ("- 'crash' (no device) = bug_report", False),
        ("- When in doubt, Y is fine — we'll catch edge cases in analysis", False),
        ("", False),
        ("ESTIMATED TIME: ~5-10 seconds per review", True),
    ]

    for i, (text, bold) in enumerate(instructions, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=12 if bold else 11)
        cell.alignment = WRAP_ALIGN


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2: Review Verification
# ══════════════════════════════════════════════════════════════════════════════

def create_verification_sheet(wb, reviews):
    ws = wb.create_sheet("Verify Reviews")
    ws.sheet_properties.tabColor = "00B050"

    # Headers
    headers = [
        "ID", "Review Text", "Rating", "LLM Label",
        "LLM Confidence", "LLM Reasoning", "Keyword Match",
        "Correct? (Y/N)", "Correct Label (if N)", "Comments"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header(ws, 1, len(headers))

    # Column widths
    widths = [6, 70, 8, 16, 14, 40, 20, 14, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data validation for Y/N column
    yn_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    yn_validation.error = "Please enter Y or N"
    yn_validation.errorTitle = "Invalid input"
    ws.add_data_validation(yn_validation)

    # Data validation for correct label column
    label_list = ",".join(LABELS)
    label_validation = DataValidation(type="list", formula1=f'"{label_list}"', allow_blank=True)
    label_validation.error = "Please select a valid label"
    label_validation.errorTitle = "Invalid label"
    ws.add_data_validation(label_validation)

    # Write review data
    for i, r in enumerate(reviews, 2):
        ws.cell(row=i, column=1, value=i - 1)  # ID
        ws.cell(row=i, column=2, value=r["text"]).alignment = WRAP_ALIGN
        ws.cell(row=i, column=3, value=r.get("rating", "")).alignment = CENTER_ALIGN
        ws.cell(row=i, column=4, value=r.get("llm_label", "")).alignment = CENTER_ALIGN
        ws.cell(row=i, column=5, value=r.get("llm_confidence", 0)).alignment = CENTER_ALIGN
        ws.cell(row=i, column=6, value=r.get("llm_reasoning", "")).alignment = WRAP_ALIGN
        ws.cell(row=i, column=7, value=", ".join(r.get("keyword_matched_categories", []))).alignment = CENTER_ALIGN

        # Yellow cells for human input
        yn_cell = ws.cell(row=i, column=8)
        yn_cell.fill = VERIFY_FILL
        yn_cell.alignment = CENTER_ALIGN
        yn_validation.add(yn_cell)

        label_cell = ws.cell(row=i, column=9)
        label_cell.fill = VERIFY_FILL
        label_cell.alignment = CENTER_ALIGN
        label_validation.add(label_cell)

        comment_cell = ws.cell(row=i, column=10)
        comment_cell.fill = VERIFY_FILL
        comment_cell.alignment = WRAP_ALIGN

        # Add borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = THIN_BORDER

    # Freeze top row
    ws.freeze_panes = "A2"

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3: Summary Statistics
# ══════════════════════════════════════════════════════════════════════════════

def create_summary_sheet(wb, reviews):
    ws = wb.create_sheet("Statistics")
    ws.sheet_properties.tabColor = "FFC000"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    label_counts = Counter(r.get("llm_label", "other") for r in reviews)
    total = len(reviews)

    ws.cell(row=1, column=1, value="LLM Label Distribution")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    headers = ["Label", "Count", "Percentage"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 3)

    for i, label in enumerate(LABELS, 4):
        count = label_counts.get(label, 0)
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=count)
        ws.cell(row=i, column=3, value=f"{count/total*100:.1f}%")

    ws.cell(row=4 + len(LABELS), column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=4 + len(LABELS), column=2, value=total).font = Font(bold=True)

    # Confidence breakdown
    row = 4 + len(LABELS) + 2
    ws.cell(row=row, column=1, value="Confidence Breakdown").font = Font(bold=True, size=14)
    row += 2

    confidences = [r.get("llm_confidence", 0) for r in reviews]
    brackets = [
        ("High (>=0.9)", sum(1 for c in confidences if c >= 0.9)),
        ("Good (0.8-0.9)", sum(1 for c in confidences if 0.8 <= c < 0.9)),
        ("Medium (0.6-0.8)", sum(1 for c in confidences if 0.6 <= c < 0.8)),
        ("Low (<0.6)", sum(1 for c in confidences if c < 0.6)),
    ]
    for label, count in brackets:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%")
        row += 1


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Generate verification spreadsheet from LLM-labeled reviews")
    parser.add_argument("--input", default="data/processed/rrgen_llm_labeled/llm_labeled_all.json",
                        help="Path to LLM-labeled JSON")
    parser.add_argument("--output", default="Synthetic_Data_Verification_RRGen.xlsx",
                        help="Output Excel file path")
    parser.add_argument("--samples", type=int, default=None,
                        help="Limit to first N reviews (default: all)")
    args = parser.parse_args()

    # Load LLM-labeled data
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: {input_path} not found.")
        print("Run llm_label_rrgen.py first.")
        sys.exit(1)

    with open(input_path) as f:
        reviews = json.load(f)

    if args.samples:
        reviews = reviews[:args.samples]

    print(f"Loaded {len(reviews):,} LLM-labeled reviews")

    # Sort by LLM label for easier reviewing
    reviews.sort(key=lambda r: (r.get("llm_label", "other"), -r.get("llm_confidence", 0)))

    # Create workbook
    wb = Workbook()
    create_instructions_sheet(wb)
    create_verification_sheet(wb, reviews)
    create_summary_sheet(wb, reviews)

    # Save
    output_path = Path(args.output)
    wb.save(output_path)
    print(f"\nSaved verification sheet: {output_path}")
    print(f"  Reviews: {len(reviews):,}")
    print(f"  Sheets: Instructions, Verify Reviews, Statistics")
    print(f"\nNext step: Have reviewers fill in the yellow columns (Y/N + correct label if N)")
    print(f"Then run: python3 scripts/ingest_verified_labels.py --input {output_path}")


if __name__ == "__main__":
    main()
