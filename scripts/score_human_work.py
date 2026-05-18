"""
Score all three completed human-work spreadsheets:

  #11 cluster_curation.xlsx   → re-cluster purity recompute
  #10 reference_specs.xlsx    → add as condition (d) human_written for Experiment 1
  #9  response_ratings.xlsx   → unblind A/B/C/D, paired Wilcoxon, human-eval Exp 2 result

Outputs land in:
  data/processed/experiments/exp1_with_human_written.json
  data/processed/experiments/exp2_human_eval.json
  data/processed/clusters_umap/cluster_curation_outcome.json
"""

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean, median, stdev

from openpyxl import load_workbook

LABELS = ["bug_report", "feature_request", "performance", "usability",
          "compatibility", "praise", "other"]


# ============================================================
# #11 — Cluster curation outcome
# ============================================================
def score_cluster_curation():
    print("="*70)
    print("#11 — CLUSTER CURATION")
    print("="*70)
    wb = load_workbook("human_work/cluster_curation.xlsx", data_only=True)
    ws = wb["Curation"]
    header = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(header) if h}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[col["cluster_id"]]: continue
        rows.append({
            "cluster_id": r[col["cluster_id"]],
            "issue_type": r[col["issue_type"]],
            "auto_name": r[col["auto_name"]],
            "verdict": str(r[col["verdict"]]).strip(),
            "notes": r[col["notes"]] or "",
        })

    actions = Counter()
    keeps, splits, merges, renames = [], [], [], []
    for r in rows:
        v = r["verdict"]
        if v == "Keep":
            actions["Keep"] += 1
            keeps.append(r)
        elif v == "Split":
            actions["Split"] += 1
            splits.append(r)
        elif v.startswith("Merge:"):
            actions["Merge"] += 1
            target = v.split(":", 1)[1].strip()
            merges.append({**r, "merge_target": target})
        elif v.startswith("Rename:"):
            actions["Rename"] += 1
            new_name = v.split(":", 1)[1].strip()
            renames.append({**r, "new_name": new_name})

    print(f"Total curated: {len(rows)}")
    print(f"  Keep:   {actions['Keep']}")
    print(f"  Split:  {actions['Split']}")
    print(f"  Merge:  {actions['Merge']}")
    print(f"  Rename: {actions['Rename']}")

    # Effective cluster count after curation:
    # Start from 194, +Split clusters add ~2x (each split creates ~2-3 new), -Merge clusters reduce
    # We treat "Keep" as 1, "Rename" as 1, "Merge" as 0 (folds into target), "Split" as 2.5 avg
    new_count = (actions["Keep"] + actions["Rename"]
                 + 2.5 * actions["Split"]
                 + 0 * actions["Merge"])
    # Plus uncurated 94 clusters (194-100 sampled)
    untouched = 194 - 100
    final_estimated_clusters = untouched + new_count
    print(f"\nCurated subset (100 of 194): post-curation effective ≈ {new_count:.0f} clusters")
    print(f"Including untouched 94: total ≈ {final_estimated_clusters:.0f} clusters in curated dataset")

    # Recompute purity proxy: Keep + Rename count as "purity 1.0", Split = 0.4 (mixed),
    # Merge = 0.5 (was duplicate). This is the "curation-aware purity"
    curation_purity_score = (
        actions["Keep"] * 1.0
        + actions["Rename"] * 1.0   # name was wrong but content was coherent
        + actions["Split"] * 0.4    # mixed themes → low purity
        + actions["Merge"] * 0.5    # duplicate → moderate
    ) / len(rows)
    print(f"\nCuration-aware purity (1.0 for Keep/Rename, 0.4 for Split, 0.5 for Merge):")
    print(f"  {curation_purity_score:.3f}  on the 100-cluster sample")

    out = {
        "n_curated": len(rows),
        "actions": dict(actions),
        "curation_purity_score": round(curation_purity_score, 3),
        "estimated_post_curation_clusters_in_sample": round(new_count, 1),
        "estimated_total_clusters": round(final_estimated_clusters, 0),
        "splits": splits,
        "merges": merges,
        "renames": renames,
    }
    Path("data/processed/clusters_umap").mkdir(parents=True, exist_ok=True)
    with open("data/processed/clusters_umap/cluster_curation_outcome.json", "w") as f:
        json.dump(out, f, indent=2)
    print(f"\nSaved data/processed/clusters_umap/cluster_curation_outcome.json")
    return out


# ============================================================
# #10 — Reference specs → Experiment 1 condition (d)
# ============================================================
def score_reference_specs():
    print("\n" + "="*70)
    print("#10 — REFERENCE SPECS (Experiment 1 condition d)")
    print("="*70)
    wb = load_workbook("human_work/reference_specs.xlsx", data_only=True)
    ws = wb["Specs"]
    header = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(header) if h}

    specs = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[col["cluster_id"]]: continue
        spec = {
            "issue_id": f"is_{r[col['cluster_id']]}",
            "cluster_id": r[col["cluster_id"]],
            "issue_type": r[col["issue_type"]],
            "title": r[col["your_title"]] or "",
            "description": r[col["your_description"]] or "",
            "severity": r[col["your_severity"]] or "P2",
            "affected_component": r[col["your_affected_component"]] or "",
            "steps_to_reproduce": (str(r[col["your_steps_to_reproduce"]]).split(" | ")
                                   if r[col["your_steps_to_reproduce"]] else None),
            "expected_behavior": r[col["your_expected_behavior"]] or None,
            "actual_behavior": r[col["your_actual_behavior"]] or None,
            "user_story": r[col["your_user_story"]] or None,
            "acceptance_criteria": (str(r[col["your_acceptance_criteria"]]).split(" | ")
                                    if r[col["your_acceptance_criteria"]] else None),
            "nfr_category": r[col["your_nfr_category"]] or None,
            "nielsen_heuristic": r[col["your_nielsen_heuristic"]] or None,
            "device_os_matrix": r[col["your_device_os_matrix"]] or None,
            "condition": "human_written",
        }
        specs.append(spec)

    # Save condition (d)
    Path("data/processed/issue_specs").mkdir(parents=True, exist_ok=True)
    with open("data/processed/issue_specs/specs_human_written.json", "w") as f:
        json.dump(specs, f, indent=2)

    # Run completeness on condition (d)
    def schema_required(it):
        base = ["title", "description", "severity", "affected_component"]
        m = {"bug_report": ["steps_to_reproduce", "expected_behavior", "actual_behavior"],
             "feature_request": ["user_story", "acceptance_criteria"],
             "performance": ["nfr_category"],
             "usability": ["nielsen_heuristic"],
             "compatibility": ["device_os_matrix"]}
        return base + m.get(it, [])

    def completeness(s):
        req = schema_required(s.get("issue_type"))
        filled = sum(1 for f in req if s.get(f) not in (None, "", [], {}))
        return filled / len(req) if req else 0.0

    comp_scores = [completeness(s) for s in specs]
    desc_lens = [len((s.get("description") or "").split()) for s in specs]
    sev_dist = Counter(s.get("severity") for s in specs)
    type_dist = Counter(s.get("issue_type") for s in specs)

    cond_d = {
        "n": len(specs),
        "completeness_ratio_mean": round(mean(comp_scores), 4),
        "completeness_ratio_median": round(median(comp_scores), 4),
        "description_length_mean_words": round(mean(desc_lens), 1),
        "description_length_median_words": round(median(desc_lens), 1),
        "severity_distribution": dict(sev_dist),
        "issue_type_distribution": dict(type_dist),
    }
    print(f"Condition (d) human_written:")
    for k, v in cond_d.items():
        print(f"  {k}: {v}")

    # Merge into Experiment 1 results
    exp1_path = Path("data/processed/experiments/exp1_results.json")
    exp1 = json.load(open(exp1_path)) if exp1_path.exists() else {}
    exp1["human_written"] = cond_d
    with open("data/processed/experiments/exp1_with_human_written.json", "w") as f:
        json.dump(exp1, f, indent=2)
    print(f"\nSaved exp1_with_human_written.json")
    return exp1


# ============================================================
# #9 — Response ratings → Experiment 2 human evaluation
# ============================================================
def score_response_ratings():
    print("\n" + "="*70)
    print("#9 — RESPONSE RATINGS (Experiment 2 human eval)")
    print("="*70)

    # Load blinding map
    blinding = json.load(open("human_work/response_ratings_blinding.json"))
    # Build lookup: (review_index, blind_letter) -> true condition
    blind_lookup = {}
    for entry in blinding:
        for letter, cond in entry["blinding"].items():
            blind_lookup[(entry["review_index"], letter)] = cond

    wb = load_workbook("human_work/response_ratings.xlsx", data_only=True)
    ws = wb["Ratings"]
    header = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(header) if h}

    # Per-condition aggregate
    by_cond = {c: {"quality": [], "specificity": [], "helpful_y": 0, "helpful_n": 0,
                   "per_issue_type": defaultdict(list)}
               for c in ["rrgen_baseline", "prompt_baseline", "reviewagent_no_spec", "reviewagent_full"]}
    # Per (review, condition) for paired tests
    paired = defaultdict(dict)  # review_idx -> {cond: {q, s, h}}

    for r in ws.iter_rows(min_row=2, values_only=True):
        rid = r[col["review_index"]]
        if rid is None: continue
        rid = int(rid)
        blind_letter = r[col["blind_id"]]
        true_cond = blind_lookup.get((rid, blind_letter))
        if true_cond is None: continue

        q = r[col["quality_1_to_5"]]
        s = r[col["specificity_1_to_5"]]
        h = r[col["helpful_y_n"]]
        issue_type = r[col["issue_type"]]

        if q is None or s is None: continue
        try:
            q = float(q); s = float(s)
        except: continue
        h_str = str(h).strip().upper() if h else ""

        by_cond[true_cond]["quality"].append(q)
        by_cond[true_cond]["specificity"].append(s)
        by_cond[true_cond]["per_issue_type"][issue_type].append((q, s))
        if h_str == "Y":
            by_cond[true_cond]["helpful_y"] += 1
        elif h_str == "N":
            by_cond[true_cond]["helpful_n"] += 1

        paired[rid][true_cond] = {"q": q, "s": s, "h": h_str}

    # Aggregate stats
    summary = {}
    for cond, vals in by_cond.items():
        n = len(vals["quality"])
        if n == 0: continue
        summary[cond] = {
            "n": n,
            "quality_mean": round(mean(vals["quality"]), 3),
            "quality_std": round(stdev(vals["quality"]), 3) if n > 1 else 0,
            "specificity_mean": round(mean(vals["specificity"]), 3),
            "specificity_std": round(stdev(vals["specificity"]), 3) if n > 1 else 0,
            "helpful_y": vals["helpful_y"],
            "helpful_n": vals["helpful_n"],
            "helpful_pct": round(100*vals["helpful_y"]/(vals["helpful_y"]+vals["helpful_n"]), 1) if (vals["helpful_y"]+vals["helpful_n"]) else 0,
        }
        # Per issue type
        per_type = {}
        for it, qs in vals["per_issue_type"].items():
            per_type[it] = {
                "n": len(qs),
                "quality_mean": round(mean([x[0] for x in qs]), 3),
                "specificity_mean": round(mean([x[1] for x in qs]), 3),
            }
        summary[cond]["per_issue_type"] = per_type

    # Print summary table
    print(f"\n{'condition':25s} {'n':>4} {'quality':>10} {'specificity':>13} {'helpful%':>10}")
    print("-"*70)
    for cond in ["rrgen_baseline", "prompt_baseline", "reviewagent_no_spec", "reviewagent_full"]:
        if cond not in summary: continue
        s = summary[cond]
        print(f"{cond:25s} {s['n']:>4} "
              f"{s['quality_mean']:>5.2f}±{s['quality_std']:>4.2f} "
              f"{s['specificity_mean']:>7.2f}±{s['specificity_std']:>4.2f} "
              f"{s['helpful_pct']:>9.1f}%")

    # Paired Wilcoxon: full vs no_spec, full vs core, full vs rrgen
    def wilcoxon(a, b):
        """Simple paired Wilcoxon implementation (signed-rank test)."""
        diffs = [(x - y) for x, y in zip(a, b) if (x - y) != 0]
        if not diffs:
            return None, None
        abs_diffs = [(abs(d), 1 if d > 0 else -1) for d in diffs]
        abs_diffs.sort(key=lambda x: x[0])
        # Rank with ties averaged
        ranks = [0.0] * len(abs_diffs)
        i = 0
        while i < len(abs_diffs):
            j = i
            while j+1 < len(abs_diffs) and abs_diffs[j+1][0] == abs_diffs[i][0]:
                j += 1
            avg_rank = (i + j) / 2 + 1
            for k in range(i, j+1):
                ranks[k] = avg_rank
            i = j+1
        W_plus = sum(rank for rank, (_, sign) in zip(ranks, abs_diffs) if sign > 0)
        n = len(diffs)
        # Approximate via normal for n > 20
        import math
        mean_W = n*(n+1)/4
        var_W = n*(n+1)*(2*n+1)/24
        if var_W == 0: return None, None
        z = (W_plus - mean_W) / math.sqrt(var_W)
        # Two-tailed p-value
        from math import erf, sqrt
        p = 2 * (1 - 0.5 * (1 + erf(abs(z) / sqrt(2))))
        return round(z, 3), round(p, 5)

    def paired_test(cond_a, cond_b, metric="q"):
        a_vals, b_vals = [], []
        for rid, conds in paired.items():
            if cond_a in conds and cond_b in conds:
                a_vals.append(conds[cond_a][metric])
                b_vals.append(conds[cond_b][metric])
        z, p = wilcoxon(a_vals, b_vals)
        return {"n_pairs": len(a_vals), "z": z, "p_value": p,
                "mean_diff": round(mean([a-b for a, b in zip(a_vals, b_vals)]), 3) if a_vals else 0}

    print("\nPaired Wilcoxon — quality scores:")
    pairs = [
        ("reviewagent_full", "reviewagent_no_spec"),
        ("reviewagent_full", "prompt_baseline"),
        ("reviewagent_full", "rrgen_baseline"),
        ("reviewagent_no_spec", "prompt_baseline"),
        ("reviewagent_no_spec", "rrgen_baseline"),
        ("prompt_baseline", "rrgen_baseline"),
    ]
    pair_results = {}
    for a, b in pairs:
        r = paired_test(a, b, "q")
        pair_results[f"{a}_vs_{b}_quality"] = r
        sig = "***" if r["p_value"] and r["p_value"] < 0.001 else ("**" if r["p_value"] and r["p_value"] < 0.01 else ("*" if r["p_value"] and r["p_value"] < 0.05 else ""))
        print(f"  {a} vs {b}: Δ={r['mean_diff']:+.2f}  p={r['p_value']}{sig}")

    print("\nPaired Wilcoxon — specificity scores:")
    for a, b in pairs:
        r = paired_test(a, b, "s")
        pair_results[f"{a}_vs_{b}_specificity"] = r
        sig = "***" if r["p_value"] and r["p_value"] < 0.001 else ("**" if r["p_value"] and r["p_value"] < 0.01 else ("*" if r["p_value"] and r["p_value"] < 0.05 else ""))
        print(f"  {a} vs {b}: Δ={r['mean_diff']:+.2f}  p={r['p_value']}{sig}")

    final = {
        "summary": summary,
        "paired_wilcoxon": pair_results,
        "n_pairs": len(paired),
    }
    Path("data/processed/experiments").mkdir(parents=True, exist_ok=True)
    with open("data/processed/experiments/exp2_human_eval.json", "w") as f:
        json.dump(final, f, indent=2)
    print(f"\nSaved data/processed/experiments/exp2_human_eval.json")
    return final


if __name__ == "__main__":
    cur = score_cluster_curation()
    refs = score_reference_specs()
    rates = score_response_ratings()
    print("\n\nDONE — three scoring scripts run.")
