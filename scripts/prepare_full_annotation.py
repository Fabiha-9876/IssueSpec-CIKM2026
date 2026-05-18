"""
Prepare FULL 215K RRGen Dataset for Manual Annotation
======================================================

Exports ALL 215,543 reviews into CSV files (one per category)
with Y/N columns for annotators.

Excel is too slow for 215K rows, so we use CSV + a small Excel summary.

Output:
    <HOME>/Desktop/Review Agent/RRGen_Annotation/
        bug_report.csv          (80,056 reviews)
        feature_request.csv     (26,286 reviews)
        performance.csv         (182 reviews)
        usability.csv           (5,001 reviews)
        compatibility.csv       (8 reviews)
        praise.csv              (57,919 reviews)
        other.csv               (46,091 reviews)
        all_reviews.csv         (215,543 reviews - single file)
        annotation_summary.xlsx (summary + instructions)
"""

import json
import csv
import sys
from pathlib import Path
from collections import Counter
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.stage1.classifier import LABELS

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Warning: openpyxl not found, skipping summary Excel")


def main():
    output_dir = Path("<HOME>/Desktop/Review Agent/RRGen_Annotation")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load data
    print("Loading labeled data...")
    with open("data/processed/rrgen_full_labeled/rrgen_full_labeled.json") as f:
        all_reviews = json.load(f)
    print(f"  Total reviews: {len(all_reviews):,}")

    # Sort by confidence (lowest first - most uncertain at top)
    all_reviews.sort(key=lambda r: (r["predicted_label"], r["confidence"]))

    # Group by label
    by_label = {}
    for r in all_reviews:
        label = r["predicted_label"]
        if label not in by_label:
            by_label[label] = []
        by_label[label].append(r)

    # CSV headers
    csv_headers = [
        "id", "text", "rating", "predicted_label", "confidence",
        "needs_hitl", "app_id",
        "correct_yn", "correct_label_if_no", "comments"
    ]

    # ── Write one CSV per category ───────────────────────────────────────
    print("\nCreating per-category CSV files...")
    for label in LABELS:
        reviews = by_label.get(label, [])
        if not reviews:
            print(f"  {label:20s}: 0 reviews (skipped)")
            continue

        # Sort by confidence (lowest first)
        reviews.sort(key=lambda r: r["confidence"])

        csv_path = output_dir / f"{label}.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(csv_headers)
            for i, r in enumerate(reviews, 1):
                writer.writerow([
                    i,
                    r["text"],
                    r.get("rating", ""),
                    r["predicted_label"],
                    r["confidence"],
                    "Yes" if r.get("needs_hitl") else "No",
                    r.get("app_id", ""),
                    "",  # correct_yn - annotator fills
                    "",  # correct_label_if_no - annotator fills
                    "",  # comments - annotator fills
                ])
        print(f"  {label:20s}: {len(reviews):6,} reviews -> {csv_path.name}")

    # ── Write single combined CSV ────────────────────────────────────────
    print("\nCreating combined CSV (all reviews)...")
    all_sorted = sorted(all_reviews, key=lambda r: r["confidence"])
    csv_path = output_dir / "all_reviews.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(csv_headers)
        for i, r in enumerate(all_sorted, 1):
            writer.writerow([
                i,
                r["text"],
                r.get("rating", ""),
                r["predicted_label"],
                r["confidence"],
                "Yes" if r.get("needs_hitl") else "No",
                r.get("app_id", ""),
                "",  # correct_yn
                "",  # correct_label_if_no
                "",  # comments
            ])
    print(f"  Saved: {csv_path.name} ({len(all_sorted):,} rows)")

    # ── Summary Excel ────────────────────────────────────────────────────
    print("\nCreating summary Excel...")
    try:
        wb = Workbook()

        # Instructions sheet
        ws = wb.active
        ws.title = "Instructions"
        ws.column_dimensions["A"].width = 90

        lines = [
            ("MANUAL ANNOTATION INSTRUCTIONS", True, 14),
            ("", False, 11),
            (f"DATASET: Full RRGen ({len(all_reviews):,} reviews)", True, 12),
            (f"DATE: {datetime.now().strftime('%Y-%m-%d')}", False, 11),
            ("", False, 11),
            ("FILES:", True, 12),
            ("  Each category has its own CSV file:", False, 11),
        ]
        for label in LABELS:
            count = len(by_label.get(label, []))
            lines.append((f"    {label}.csv — {count:,} reviews", False, 11))
        lines += [
            ("    all_reviews.csv — all 215K reviews in one file", False, 11),
            ("", False, 11),
            ("HOW TO ANNOTATE:", True, 12),
            ("  1. Open any CSV file in Excel/Google Sheets", False, 11),
            ("  2. Reviews are sorted by confidence (lowest first = most uncertain)", False, 11),
            ("  3. For each review, fill these columns:", False, 11),
            ("     correct_yn: Type Y if predicted label is correct, N if wrong", False, 11),
            ("     correct_label_if_no: If you typed N, write the correct label here", False, 11),
            ("     comments: Optional notes for ambiguous cases", False, 11),
            ("", False, 11),
            ("CATEGORY DEFINITIONS:", True, 12),
            ("  bug_report      — Crashes, errors, broken features, malfunctions", False, 11),
            ("  feature_request  — Requests for new features or improvements", False, 11),
            ("  performance      — Speed, battery, memory, lag, loading times", False, 11),
            ("  usability        — Confusing UI, hard to use, poor navigation", False, 11),
            ("  compatibility    — Device-specific or OS-specific issues", False, 11),
            ("  praise           — Positive feedback, compliments", False, 11),
            ("  other            — Doesn't fit any category above", False, 11),
            ("", False, 11),
            ("LABELING RULES:", True, 12),
            ("  'slow' / 'lag' / 'battery drain' → performance (NOT bug_report)", False, 11),
            ("  'crash on my Samsung' → compatibility (device-specific issue)", False, 11),
            ("  'crash' / 'error' (no device) → bug_report", False, 11),
            ("  'good app' / 'love it' → praise", False, 11),
            ("  'please add dark mode' → feature_request", False, 11),
            ("  'confusing menu' / 'hard to navigate' → usability", False, 11),
            ("", False, 11),
            ("PRIORITY:", True, 12),
            ("  Start with performance.csv (182 reviews) and compatibility.csv (8 reviews)", False, 11),
            ("  These are the rarest categories and need the most correction.", False, 11),
            ("  Then move to other categories.", False, 11),
        ]

        for i, (text, bold, size) in enumerate(lines, 1):
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(bold=bold, size=size)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Statistics sheet
        ws2 = wb.create_sheet("Statistics")
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 15
        ws2.column_dimensions["C"].width = 15
        ws2.column_dimensions["D"].width = 20

        ws2.cell(row=1, column=1, value="Label Distribution").font = Font(bold=True, size=14)

        headers = ["Label", "Count", "Percentage", "CSV File"]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

        for i, label in enumerate(LABELS, 4):
            count = len(by_label.get(label, []))
            ws2.cell(row=i, column=1, value=label)
            ws2.cell(row=i, column=2, value=count)
            ws2.cell(row=i, column=3, value=f"{count/len(all_reviews)*100:.1f}%")
            ws2.cell(row=i, column=4, value=f"{label}.csv")

        row = 4 + len(LABELS)
        ws2.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws2.cell(row=row, column=2, value=len(all_reviews)).font = Font(bold=True)

        # Confidence stats
        row += 2
        ws2.cell(row=row, column=1, value="Confidence Breakdown").font = Font(bold=True, size=14)
        row += 2
        confs = [r["confidence"] for r in all_reviews]
        brackets = [
            ("High (>=0.9)", sum(1 for c in confs if c >= 0.9)),
            ("Good (0.8-0.9)", sum(1 for c in confs if 0.8 <= c < 0.9)),
            ("Medium (0.6-0.8)", sum(1 for c in confs if 0.6 <= c < 0.8)),
            ("Low (0.4-0.6)", sum(1 for c in confs if 0.4 <= c < 0.6)),
            ("Very Low (<0.4)", sum(1 for c in confs if c < 0.4)),
        ]
        for label, count in brackets:
            ws2.cell(row=row, column=1, value=label)
            ws2.cell(row=row, column=2, value=count)
            ws2.cell(row=row, column=3, value=f"{count/len(confs)*100:.1f}%")
            row += 1

        xlsx_path = output_dir / "annotation_summary.xlsx"
        wb.save(xlsx_path)
        print(f"  Saved: {xlsx_path.name}")
    except Exception as e:
        print(f"  Excel creation failed: {e}")

    # ── Final summary ────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"DONE! Full annotation dataset ready.")
    print(f"{'='*60}")
    print(f"\n  Location: {output_dir}")
    print(f"\n  Files:")
    for label in LABELS:
        count = len(by_label.get(label, []))
        if count > 0:
            print(f"    {label}.csv — {count:,} reviews")
    print(f"    all_reviews.csv — {len(all_reviews):,} reviews (combined)")
    print(f"    annotation_summary.xlsx — instructions + statistics")
    print(f"\n  START WITH:")
    print(f"    1. performance.csv (182 reviews) — most important to verify")
    print(f"    2. compatibility.csv (8 reviews) — most important to verify")
    print(f"    3. Then other categories")


if __name__ == "__main__":
    main()
