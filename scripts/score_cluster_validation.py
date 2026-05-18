"""
Compute cluster-purity statistics from the filled-in validation spreadsheet.

Run this AFTER you've filled in the 'coherent_yn' column in
data/processed/clusters_umap/cluster_validation.xlsx.

Output:
    Console summary:
      - Overall purity (Y / total)
      - Per-class purity
      - Treating P (partial) as 0.5
      - List of failed (N) clusters with notes
    data/processed/clusters_umap/cluster_validation_score.json
"""

import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

IN_PATH  = Path("data/processed/clusters_umap/cluster_validation.xlsx")
OUT_JSON = Path("data/processed/clusters_umap/cluster_validation_score.json")


def main():
    if not IN_PATH.exists():
        print(f"Validation file not found: {IN_PATH}")
        print("Run scripts/build_cluster_validation_sheet.py first.")
        return

    wb = load_workbook(IN_PATH, data_only=True)
    ws = wb["Cluster Validation"]

    # Header
    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append(row)

    n_total = len(rows)
    if n_total == 0:
        print("No clusters in sheet.")
        return

    # Aggregate
    overall_counts = Counter()
    per_class_counts = {}  # issue_type -> Counter
    failed = []
    unfilled = 0

    for r in rows:
        cluster_id = r[col["cluster_id"]]
        issue_type = r[col["issue_type"]]
        verdict_raw = r[col["coherent_yn"]]
        notes = r[col["notes"]] or ""

        if verdict_raw is None or str(verdict_raw).strip() == "":
            unfilled += 1
            continue

        v = str(verdict_raw).strip().upper()
        if v not in ("Y", "N", "P"):
            unfilled += 1
            continue

        overall_counts[v] += 1
        per_class_counts.setdefault(issue_type, Counter())[v] += 1

        if v in ("N", "P"):
            failed.append({
                "cluster_id": cluster_id,
                "issue_type": issue_type,
                "auto_name": r[col["auto_name"]],
                "verdict": v,
                "notes": notes,
            })

    n_judged = sum(overall_counts.values())

    def purity(counts):
        if not counts:
            return None
        n = sum(counts.values())
        # Y=1, P=0.5, N=0
        score = (counts.get("Y", 0) + 0.5 * counts.get("P", 0)) / n
        return round(score, 3), counts.get("Y", 0), counts.get("P", 0), counts.get("N", 0), n

    print("=" * 70)
    print("CLUSTER VALIDATION RESULTS")
    print("=" * 70)
    print(f"Total clusters in sheet: {n_total}")
    print(f"Judged so far:           {n_judged}")
    print(f"Unfilled:                {unfilled}")
    if n_judged == 0:
        print("\nNo verdicts yet — fill in the 'coherent_yn' column and rerun.")
        return

    p, y, pp, nn, n = purity(overall_counts)
    print(f"\nOVERALL PURITY: {p:.3f}  (Y={y}  P={pp}  N={nn}  n={n})")

    print(f"\nPer-class purity:")
    print(f"  {'class':18s} {'purity':>7s} {'Y':>4s} {'P':>4s} {'N':>4s} {'n':>4s}")
    per_class_summary = {}
    for cls, cnts in sorted(per_class_counts.items()):
        result = purity(cnts)
        if result is None:
            continue
        p, y, pp, nn, n = result
        per_class_summary[cls] = {"purity": p, "Y": y, "P": pp, "N": nn, "n": n}
        print(f"  {cls:18s} {p:>7.3f} {y:>4d} {pp:>4d} {nn:>4d} {n:>4d}")

    if failed:
        print(f"\nFailed/partial clusters ({len(failed)}):")
        for f in failed:
            note_str = f" — {f['notes']}" if f['notes'] else ""
            print(f"  [{f['verdict']}] {f['cluster_id']:10s} {f['issue_type']:18s}  {f['auto_name']}{note_str}")

    # Save summary
    out = {
        "n_total": n_total,
        "n_judged": n_judged,
        "n_unfilled": unfilled,
        "overall_purity": purity(overall_counts)[0] if n_judged else None,
        "overall_counts": dict(overall_counts),
        "per_class": per_class_summary,
        "failed_clusters": failed,
    }
    with open(OUT_JSON, "w") as f:
        json.dump(out, f, indent=2)
    print(f"\nSaved: {OUT_JSON}")


if __name__ == "__main__":
    main()
