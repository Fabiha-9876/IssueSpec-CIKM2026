"""
Ingest Human-Verified Labels Back into Training Dataset
========================================================

Reads the completed verification spreadsheet and:
1. Accepts LLM labels where human said Y
2. Uses human-corrected labels where human said N
3. Skips unreviewed rows
4. Merges with existing training data (MAALEJ + Synthetic)
5. Creates held-out test set (SEPARATE — never used in training)
6. Outputs final extended training dataset

Usage:
    python3 scripts/ingest_verified_labels.py
    python3 scripts/ingest_verified_labels.py --input Synthetic_Data_Verification_RRGen.xlsx
    python3 scripts/ingest_verified_labels.py --test-split 0.15  # 15% held-out test

Output:
    data/processed/extended/
        train_extended.json       — training set (MAALEJ + Synthetic + Verified RRGen)
        test_holdout.json         — held-out test set (NEVER use for training)
        val_split.json            — validation set (for eval during training)
        ingestion_stats.json      — summary statistics
        label_corrections.json    — records of what humans corrected (for analysis)
"""

import json
import sys
import argparse
from pathlib import Path
from collections import Counter
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

import numpy as np
from sklearn.model_selection import train_test_split

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.stage1.classifier import LABELS


def read_verification_sheet(xlsx_path):
    """Read completed verification spreadsheet and extract verified labels."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Verify Reviews"]

    verified = []
    skipped = 0
    corrections = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        cells = [c.value for c in row]
        if not cells[0] or not cells[1]:  # skip empty rows
            continue

        review_id = cells[0]
        text = cells[1]
        rating = cells[2] or 0
        llm_label = cells[3] or "other"
        llm_confidence = cells[4] or 0
        llm_reasoning = cells[5] or ""
        keyword_match = cells[6] or ""
        correct_yn = str(cells[7] or "").strip().upper()
        correct_label = str(cells[8] or "").strip() if cells[8] else None
        comment = cells[9] or ""

        # Skip unreviewed rows
        if correct_yn not in ("Y", "N"):
            skipped += 1
            continue

        if correct_yn == "Y":
            final_label = llm_label
        else:
            # Human corrected
            if correct_label and correct_label in LABELS:
                final_label = correct_label
            else:
                # N but no correct label provided — skip
                skipped += 1
                continue

            corrections.append({
                "text": text[:100],
                "llm_label": llm_label,
                "correct_label": final_label,
                "llm_confidence": llm_confidence,
                "comment": comment,
            })

        verified.append({
            "text": text,
            "rating": int(rating),
            "labels": [final_label],
            "confidence": 1.0,  # human-verified
            "source": "rrgen_verified",
            "llm_original_label": llm_label,
            "llm_confidence": llm_confidence,
            "human_verified": True,
            "human_corrected": correct_yn == "N",
        })

    return verified, corrections, skipped


def load_existing_training_data():
    """Load MAALEJ + Synthetic training data."""
    existing = []

    # Synthetic reviews
    synthetic_path = Path("data/raw/sample_reviews.json")
    if synthetic_path.exists():
        with open(synthetic_path) as f:
            synthetic = json.load(f)
        for r in synthetic:
            existing.append({
                "text": r["text"],
                "rating": r.get("rating", 0),
                "labels": r["labels"] if isinstance(r["labels"], list) else [r["labels"]],
                "confidence": 1.0,
                "source": "synthetic",
            })
        print(f"  Loaded synthetic: {len(synthetic):,}")

    # MAALEJ labeled reviews
    maalej_path = Path("data/raw/maalej/maalej_labeled.json")
    if maalej_path.exists():
        with open(maalej_path) as f:
            maalej = json.load(f)
        for r in maalej:
            existing.append({
                "text": r["text"],
                "rating": r.get("rating", 0),
                "labels": r["labels"] if isinstance(r["labels"], list) else [r["labels"]],
                "confidence": 1.0,
                "source": "maalej",
            })
        print(f"  Loaded MAALEJ: {len(maalej):,}")
    else:
        # Try labeled_reviews.json
        labeled_path = Path("data/processed/labeled_reviews.json")
        if labeled_path.exists():
            with open(labeled_path) as f:
                labeled = json.load(f)
            for r in labeled:
                existing.append({
                    "text": r["text"],
                    "rating": r.get("rating", 0),
                    "labels": r["labels"] if isinstance(r["labels"], list) else [r["labels"]],
                    "confidence": 1.0,
                    "source": r.get("source", "maalej"),
                })
            print(f"  Loaded labeled_reviews: {len(labeled):,}")

    return existing


def main():
    parser = argparse.ArgumentParser(description="Ingest verified labels into training dataset")
    parser.add_argument("--input", default="Synthetic_Data_Verification_RRGen.xlsx",
                        help="Path to completed verification spreadsheet")
    parser.add_argument("--test-split", type=float, default=0.15,
                        help="Fraction for held-out test set (default: 0.15)")
    parser.add_argument("--val-split", type=float, default=0.10,
                        help="Fraction for validation set (default: 0.10)")
    parser.add_argument("--seed", type=int, default=42, help="Random seed")
    args = parser.parse_args()

    np.random.seed(args.seed)

    output_dir = Path("data/processed/extended")
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── Step 1: Read verified labels ─────────────────────────────────────
    print(f"Reading verification sheet: {args.input}")
    xlsx_path = Path(args.input)
    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found.")
        sys.exit(1)

    verified, corrections, skipped = read_verification_sheet(xlsx_path)
    print(f"  Verified reviews: {len(verified):,}")
    print(f"  Human corrections: {len(corrections):,}")
    print(f"  Skipped (unreviewed): {skipped:,}")

    if not verified:
        print("Error: No verified reviews found. Did reviewers fill in the Y/N column?")
        sys.exit(1)

    # ── Step 2: Load existing training data ──────────────────────────────
    print("\nLoading existing training data...")
    existing = load_existing_training_data()
    print(f"  Total existing: {len(existing):,}")

    # ── Step 3: Merge (deduplicate by text) ──────────────────────────────
    print("\nMerging datasets...")
    seen_texts = set()
    all_data = []

    for r in existing + verified:
        if r["text"] not in seen_texts:
            seen_texts.add(r["text"])
            all_data.append(r)

    print(f"  Total after dedup: {len(all_data):,}")

    # ── Step 4: Create stratified train/val/test split ───────────────────
    print(f"\nCreating splits (test={args.test_split}, val={args.val_split})...")

    # Get primary label for stratification
    primary_labels = []
    for r in all_data:
        lbl = r["labels"][0] if r["labels"] else "other"
        primary_labels.append(lbl)

    # First split: separate test set
    train_val_data, test_data, train_val_labels, test_labels = train_test_split(
        all_data, primary_labels,
        test_size=args.test_split,
        random_state=args.seed,
        stratify=primary_labels,
    )

    # Second split: separate validation from training
    val_fraction = args.val_split / (1.0 - args.test_split)
    train_data, val_data, _, _ = train_test_split(
        train_val_data, train_val_labels,
        test_size=val_fraction,
        random_state=args.seed,
        stratify=train_val_labels,
    )

    print(f"  Train: {len(train_data):,}")
    print(f"  Val:   {len(val_data):,}")
    print(f"  Test:  {len(test_data):,}  *** HELD OUT — DO NOT USE FOR TRAINING ***")

    # ── Step 5: Distribution check ───────────────────────────────────────
    print("\nLabel distribution:")
    for split_name, split_data in [("Train", train_data), ("Val", val_data), ("Test", test_data)]:
        counts = Counter(r["labels"][0] for r in split_data)
        print(f"\n  {split_name} ({len(split_data):,}):")
        for label in LABELS:
            count = counts.get(label, 0)
            pct = count / len(split_data) * 100
            print(f"    {label:20s}: {count:5d} ({pct:5.1f}%)")

    source_counts = Counter(r.get("source", "unknown") for r in all_data)
    print(f"\n  Source distribution (all data):")
    for src, count in source_counts.most_common():
        print(f"    {src:20s}: {count:,}")

    # ── Step 6: Save ─────────────────────────────────────────────────────
    train_path = output_dir / "train_extended.json"
    with open(train_path, "w") as f:
        json.dump(train_data, f, indent=2)
    print(f"\n  Saved train: {train_path}")

    val_path = output_dir / "val_split.json"
    with open(val_path, "w") as f:
        json.dump(val_data, f, indent=2)
    print(f"  Saved val:   {val_path}")

    test_path = output_dir / "test_holdout.json"
    with open(test_path, "w") as f:
        json.dump(test_data, f, indent=2)
    print(f"  Saved test:  {test_path}  *** DO NOT USE FOR TRAINING ***")

    # Save corrections for analysis
    corr_path = output_dir / "label_corrections.json"
    with open(corr_path, "w") as f:
        json.dump(corrections, f, indent=2)
    print(f"  Saved corrections: {corr_path}")

    # Stats
    stats = {
        "timestamp": datetime.now().isoformat(),
        "input_file": str(xlsx_path),
        "verified_reviews": len(verified),
        "human_corrections": len(corrections),
        "correction_rate": round(len(corrections) / len(verified) * 100, 2) if verified else 0,
        "skipped_unreviewed": skipped,
        "existing_data": len(existing),
        "total_after_dedup": len(all_data),
        "splits": {
            "train": len(train_data),
            "val": len(val_data),
            "test": len(test_data),
        },
        "test_split_ratio": args.test_split,
        "val_split_ratio": args.val_split,
        "label_distribution": {
            "train": dict(Counter(r["labels"][0] for r in train_data)),
            "val": dict(Counter(r["labels"][0] for r in val_data)),
            "test": dict(Counter(r["labels"][0] for r in test_data)),
        },
        "source_distribution": dict(source_counts),
    }
    stats_path = output_dir / "ingestion_stats.json"
    with open(stats_path, "w") as f:
        json.dump(stats, f, indent=2)
    print(f"  Saved stats: {stats_path}")

    print(f"\n{'='*60}")
    print(f"DONE! Extended dataset ready.")
    print(f"  Train with: data/processed/extended/train_extended.json")
    print(f"  Evaluate on: data/processed/extended/val_split.json")
    print(f"  Final test: data/processed/extended/test_holdout.json")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
