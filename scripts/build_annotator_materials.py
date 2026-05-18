"""
Build materials for the 3-annotator gold-standard verification (Tasks 2.8 + 2.9).

Outputs:
    annotator_materials/
        00_README.md                 instructions to share with each volunteer
        calibration_set.xlsx         20 reviews × all annotators (calibration round)
        annotator_A.xlsx             500 reviews, blank annotation columns
        annotator_B.xlsx             same 500 reviews
        annotator_C.xlsx             same 500 reviews
        master_key.json              ground-truth-style metadata (V5 label, V2 label,
                                     corrected_v2 label, app_id, confidence) — DO NOT
                                     SHARE WITH ANNOTATORS. Used for scoring later.

Sampling per ANNOTATION_PROTOCOL.md:
    - Stratified by V5 predicted label (~70 per class for 7 classes ≈ 490)
    - Within each class, sample across confidence bands (high / medium / low)
    - Different apps where possible (no app dominates >10%)
    - Seed: 42 (reproducible)

Calibration set is hand-picked to span:
    - Clear examples of all 7 classes
    - Boundary cases the protocol's decision rules call out
    - A couple of intentionally ambiguous cases
"""

import json
import random
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

LABELS = ["bug_report", "feature_request", "performance", "usability",
          "compatibility", "praise", "other"]

OUT_DIR = Path("annotator_materials")
INPUT_PATH = Path("data/processed/rrgen_v5_relabeled/rrgen_v5_relabeled.json")

PER_CLASS = 70   # 7 classes × 70 = 490 (target ~500 per protocol)
PER_APP_CAP = 50
SEED = 42

CALIBRATION_SIZE = 20


def load_dataset():
    print(f"Loading {INPUT_PATH}")
    with open(INPUT_PATH) as f:
        rows = json.load(f)
    print(f"  {len(rows):,} rows")
    return rows


def confidence_band(c):
    if c is None:
        return "unknown"
    if c >= 0.85:
        return "high"
    if c >= 0.65:
        return "medium"
    return "low"


def stratified_sample(rows):
    """Stratify by V5 label × confidence band × app spread."""
    rng = random.Random(SEED)
    by_class = defaultdict(list)
    for i, r in enumerate(rows):
        by_class[r["v5_label"]].append((i, r))

    sample = []
    for cls in LABELS:
        pool = by_class.get(cls, [])
        rng.shuffle(pool)

        # Bin by confidence band
        by_band = defaultdict(list)
        for i, r in pool:
            by_band[confidence_band(r.get("v5_confidence"))].append((i, r))

        # Even allocation across bands; if a band is short, over-fill from another
        per_band_target = PER_CLASS // 3
        picked = []
        for band in ["high", "medium", "low"]:
            picked.extend(by_band[band][:per_band_target])

        # Top up to PER_CLASS if any band was short
        if len(picked) < PER_CLASS:
            picked_ids = {i for i, _ in picked}
            extras = [(i, r) for i, r in pool if i not in picked_ids]
            picked.extend(extras[: PER_CLASS - len(picked)])

        # Limit per-app dominance
        app_counter = Counter()
        capped = []
        for i, r in picked:
            app = r.get("app_id", "unknown")
            if app_counter[app] < PER_APP_CAP:
                capped.append((i, r))
                app_counter[app] += 1
        sample.extend(capped[:PER_CLASS])

    rng.shuffle(sample)
    return sample


def pick_calibration(rows):
    """Hand-curated 20 reviews spanning all classes + boundary cases."""
    by_class = defaultdict(list)
    for i, r in enumerate(rows):
        if 30 <= len(r["text"]) <= 200:  # readable length
            by_class[r["v5_label"]].append((i, r))

    rng = random.Random(SEED + 1)
    cal = []
    # 2 clear examples per class = 14
    for cls in LABELS:
        pool = by_class[cls]
        rng.shuffle(pool)
        # Prefer high-confidence "easy" examples for calibration
        easy = [(i, r) for i, r in pool if r.get("v5_confidence", 0) >= 0.85]
        chosen = easy[:2] if len(easy) >= 2 else pool[:2]
        cal.extend(chosen)

    # 6 boundary cases — search by patterns
    boundary_patterns = [
        ("slow", "performance vs bug_report decision"),
        ("samsung", "compatibility vs bug_report (device-specific)"),
        ("crash", "ambiguous — could be bug_report or compatibility"),
        ("would be nice", "feature_request phrased as compliment"),
        ("hard to find", "usability vs feature_request"),
        ("ads", "usability vs other"),
    ]
    for keyword, _ in boundary_patterns:
        for i, r in enumerate(rows):
            if i in {x[0] for x in cal}:
                continue
            t = r["text"].lower()
            if keyword in t and 30 <= len(t) <= 200 and r.get("v5_confidence", 0) >= 0.6:
                cal.append((i, r))
                break

    rng.shuffle(cal)
    return cal[:CALIBRATION_SIZE]


def build_workbook(records, title, include_predicted_label=True, header_note=""):
    """Build an .xlsx with annotation columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Verify Reviews"

    # Instructions sheet
    inst = wb.create_sheet("Instructions", 0)
    inst_lines = [
        "REVIEW VERIFICATION TASK",
        "",
        f"Total reviews to verify: {len(records)}",
        f"Estimated time: ~5-10 seconds per review",
        "",
        "What to do:",
        "  1. Read the review text in column 'review_text'.",
        "  2. Look at column 'predicted_label' — that's the AI's guess.",
        "  3. In column 'correct_yn', type Y if the AI label is correct, N if wrong.",
        "  4. If N, type the correct label in 'correct_label_if_no'.",
        "  5. Optional: add a one-line note in 'comments' for ambiguous cases.",
        "",
        "Allowed labels:",
        "  - bug_report      Crashes, errors, broken features",
        "  - feature_request Requests for new features or improvements",
        "  - performance     Speed, battery, memory, lag, loading times",
        "  - usability       Confusing UI, hard to use, poor navigation",
        "  - compatibility   Device-specific or OS-specific issues",
        "  - praise          Positive feedback, compliments",
        "  - other           Doesn't fit any above category",
        "",
        "Decision rules (important — read before starting):",
        "  - 'slow', 'lag' is performance, NOT bug_report.",
        "  - 'crashes on my Samsung' is compatibility (device-specific).",
        "  - 'crash' alone (no device) is bug_report.",
        "  - 'would be nice if X' or 'please add X' is feature_request.",
        "  - 'hard to find the X button' is usability, not bug_report.",
        "  - When in doubt, mark Y if the label is reasonable; mark N only if",
        "    you're confident the label is wrong.",
        "",
        "Save the file and send it back when done. Resume any time — your fills",
        "are preserved in the cells.",
        "",
    ]
    if header_note:
        inst_lines = [header_note, ""] + inst_lines
    for i, line in enumerate(inst_lines, 1):
        inst.cell(row=i, column=1, value=line)
        if i == 1:
            inst.cell(row=i, column=1).font = Font(bold=True, size=14)
    inst.column_dimensions["A"].width = 90

    # Verify sheet header
    if include_predicted_label:
        headers = ["row_id", "review_text", "rating", "app_id",
                   "predicted_label", "correct_yn", "correct_label_if_no", "comments"]
    else:
        headers = ["row_id", "review_text", "rating", "app_id",
                   "your_label", "comments"]

    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for ri, (idx, r) in enumerate(records, start=2):
        if include_predicted_label:
            row = [idx, r["text"], r.get("rating"), r.get("app_id"),
                   r["v5_label"], "", "", ""]
        else:
            row = [idx, r["text"], r.get("rating"), r.get("app_id"), "", ""]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border

    widths = {
        "row_id": 8, "review_text": 60, "rating": 7, "app_id": 22,
        "predicted_label": 16, "your_label": 16,
        "correct_yn": 12, "correct_label_if_no": 18, "comments": 30,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 15)
    ws.freeze_panes = "B2"
    return wb


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = load_dataset()

    # Calibration (20 reviews) — same for all annotators
    print("\nBuilding calibration set...")
    cal = pick_calibration(rows)
    print(f"  picked {len(cal)} calibration reviews")
    cal_dist = Counter(r["v5_label"] for _, r in cal)
    for lbl, n in cal_dist.items():
        print(f"    {lbl:18s} {n}")
    wb = build_workbook(cal, "Calibration",
                        header_note="CALIBRATION ROUND — 20 reviews. "
                                    "All 3 annotators do the SAME 20 first to align understanding.")
    wb.save(OUT_DIR / "calibration_set.xlsx")
    print(f"  saved {OUT_DIR/'calibration_set.xlsx'}")

    # Main task — 500 stratified reviews
    print("\nBuilding main verification sample...")
    sample = stratified_sample(rows)
    print(f"  picked {len(sample)} reviews")
    dist = Counter(r["v5_label"] for _, r in sample)
    print("  per-class distribution:")
    for lbl in LABELS:
        print(f"    {lbl:18s} {dist.get(lbl,0):>3}")
    apps = Counter(r.get("app_id") for _, r in sample)
    print(f"  unique apps: {len(apps)}, top app: {apps.most_common(1)[0]}")

    # Save the same set to 3 separate annotator workbooks
    for ann in ["A", "B", "C"]:
        wb = build_workbook(sample, f"Annotator {ann}",
                            header_note=f"VERIFICATION TASK — Annotator {ann}. "
                                        "Do NOT consult other annotators while filling this in.")
        wb.save(OUT_DIR / f"annotator_{ann}.xlsx")
        print(f"  saved {OUT_DIR/f'annotator_{ann}.xlsx'}")

    # Master key — Anonymous keeps this, doesn't share with annotators
    master = {
        "calibration_indices": [idx for idx, _ in cal],
        "calibration_v5_labels": {idx: r["v5_label"] for idx, r in cal},
        "main_indices": [idx for idx, _ in sample],
        "main_v5_labels": {idx: r["v5_label"] for idx, r in sample},
        "main_v2_labels": {idx: r["v2_label"] for idx, r in sample},
        "main_corrected_v2_labels": {idx: r["corrected_v2_label"] for idx, r in sample},
        "main_v5_confidences": {idx: r.get("v5_confidence") for idx, r in sample},
        "stratification": {"seed": SEED, "per_class": PER_CLASS, "per_app_cap": PER_APP_CAP},
        "per_class_distribution": dict(dist),
    }
    with open(OUT_DIR / "master_key.json", "w") as f:
        json.dump(master, f, indent=2)
    print(f"  saved {OUT_DIR/'master_key.json'}  (DO NOT SHARE with annotators)")

    # README to share with volunteers
    readme = f"""# Volunteer Verification Task — README

Hi! Thanks for helping with this annotation task.

## What's in this folder

- `calibration_set.xlsx` — 20 reviews. Do these FIRST. We'll discuss disagreements
  before moving on, so we're all aligned on the categories.
- `annotator_X.xlsx` — your assigned main task. The 'X' will be A, B, or C — Anonymous
  will tell you which one is yours.
- This README.

## Task in one paragraph

You'll see ~500 mobile-app reviews. Each one already has a **predicted label**
from our AI classifier. Your job is to read each review and decide whether the
AI's label is correct (Y) or wrong (N). If wrong, write what the correct label
should be.

## Steps

1. Open `calibration_set.xlsx`. Read the "Instructions" sheet first.
2. Fill in the 20 calibration reviews (Y/N for each).
3. Send the calibration sheet back to Anonymous. After we discuss any
   disagreements, you'll be cleared to start the main task.
4. Open your `annotator_X.xlsx`. Same format, ~500 reviews.
5. Estimated time: ~3 hours total. You can do it in chunks — just save the file.

## Important rules

- Do **not** look at what other annotators are filling in. We're measuring
  inter-annotator agreement, so independent judgments are essential.
- Don't worry about being "right" — we adjudicate disagreements at the end.
- For ambiguous reviews, mark Y if the label is reasonable; only mark N if
  you're confident it's wrong.

## Categories (cheat sheet)

| label | example |
|---|---|
| bug_report | "App keeps crashing when I open it" |
| feature_request | "Please add a dark mode" |
| performance | "Super slow on my phone, takes forever to load" |
| usability | "Hard to find the settings menu" |
| compatibility | "Doesn't work on Samsung Galaxy S22" |
| praise | "Best app ever, love it!" |
| other | "Hi, just downloaded this" |

## Decision rules (read these — they prevent the most common mistakes)

1. **slow / lag** → `performance`, NOT bug_report. (App is degraded, not broken.)
2. **crash on my Samsung** → `compatibility` (device-specific). **crash** alone → `bug_report`.
3. **"would be nice if X"** → `feature_request`, even if it sounds like a complaint.
4. **"hard to find the X button"** → `usability`, not bug_report.
5. **Multi-aspect reviews** → pick the *primary* category. If genuinely two-headed,
   mark Y if the AI picked one of them.
6. **Spam / non-English / nonsense** → `other`.

## Compensation

Per the project agreement, you'll receive co-authorship or acknowledgment on
the resulting paper, depending on contribution scale. Anonymous will discuss
specifics.

## Contact

Questions? Email Anonymous at anonymous@example.com.

## Once you're done

Save the file (no need to rename) and email it back to Anonymous.

Thank you!
"""
    with open(OUT_DIR / "00_README.md", "w") as f:
        f.write(readme)
    print(f"  saved {OUT_DIR/'00_README.md'}")

    print(f"\nAll volunteer materials in: {OUT_DIR}/")


if __name__ == "__main__":
    main()
