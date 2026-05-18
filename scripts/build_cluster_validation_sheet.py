"""
Build a cluster-validation spreadsheet for Task 2 (~1h human work).

Picks ~50 clusters using a balanced strategy (top-N per issue type) and
emits an .xlsx with everything you need to mark coherent Y/N for each.

Usage:
    python3 scripts/build_cluster_validation_sheet.py

Output:
    data/processed/clusters_umap/cluster_validation.xlsx

Columns:
    cluster_id, issue_type, auto_name, review_count, top_aspects,
    sample_1..sample_5, coherent_yn, notes

The file opens cleanly in Apple Numbers or Excel. Fill in Y/N for each
cluster. Resume any time — your progress is preserved in the cells.
"""

import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CLUSTERS_FULL = Path("data/processed/clusters_umap/clusters_named.json")
OUT_PATH      = Path("data/processed/clusters_umap/cluster_validation.xlsx")

# Balanced sampling per issue type
PER_TYPE_COUNT = {
    "bug_report":      15,
    "feature_request": 15,
    "performance":     10,
    "usability":       6,
    "compatibility":   4,   # all of them (only 4 exist)
}
SAMPLES_PER_CLUSTER = 5


def main():
    print(f"Loading: {CLUSTERS_FULL}")
    with open(CLUSTERS_FULL) as f:
        clusters = json.load(f)
    print(f"  {len(clusters)} total clusters")

    # Group by issue type, sort by size desc, take top N per type
    by_type = {}
    for c in clusters:
        by_type.setdefault(c["issue_type"], []).append(c)
    for t in by_type:
        by_type[t].sort(key=lambda c: -c["review_count"])

    selected = []
    for t, n in PER_TYPE_COUNT.items():
        selected.extend(by_type.get(t, [])[:n])

    print(f"\nSelected {len(selected)} clusters for validation:")
    for t, n in PER_TYPE_COUNT.items():
        actual = sum(1 for s in selected if s["issue_type"] == t)
        print(f"  {t:18s} {actual} (target {n})")

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cluster Validation"

    # Instructions sheet
    inst = wb.create_sheet("Instructions", 0)
    inst_lines = [
        "CLUSTER VALIDATION TASK",
        "",
        "Goal: For each of the 50 clusters below, judge whether the cluster's",
        "auto-generated name and the 5 sample reviews are about the same theme.",
        "",
        "How to use:",
        "  1. Open the 'Cluster Validation' sheet.",
        "  2. For each cluster, read the auto_name and the 5 sample_* reviews.",
        "  3. Fill the 'coherent_yn' column:",
        "        Y = the 5 reviews are clearly about the same issue/theme",
        "        N = the reviews are unrelated or the auto_name is misleading",
        "        P = partial — most are coherent but 1-2 don't fit",
        "  4. Optional: add a one-line note in 'notes' if N or P.",
        "",
        "Estimated time: ~60 seconds per cluster, ~1 hour total.",
        "",
        "What this gives the paper:",
        "  Cluster purity rate = (# coherent clusters) / 50",
        "  Per-class purity rates",
        "  Examples of failure modes (the 'N' rows)",
        "",
        "Resume anytime — close the file, your fills are saved.",
        "",
        f"Total clusters in this sheet: {len(selected)}",
    ]
    for i, line in enumerate(inst_lines, 1):
        inst.cell(row=i, column=1, value=line)
        if i == 1:
            inst.cell(row=i, column=1).font = Font(bold=True, size=14)
    inst.column_dimensions["A"].width = 90

    # Validation sheet header
    headers = (
        ["cluster_id", "issue_type", "auto_name", "review_count", "top_aspects"]
        + [f"sample_{i+1}" for i in range(SAMPLES_PER_CLUSTER)]
        + ["coherent_yn", "notes"]
    )
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Color rows by issue type
    type_color = {
        "bug_report":      "FCE4E4",  # light red
        "feature_request": "E1F0FF",  # light blue
        "performance":     "FFF2CC",  # light yellow
        "usability":       "E2EFDA",  # light green
        "compatibility":   "EAD8F1",  # light purple
    }

    rng = random.Random(42)
    for row_i, c in enumerate(selected, start=2):
        # Pick 5 sample reviews — prefer the auto-picked representative reviews,
        # then add random extras from the cluster
        reps = list(c.get("representative_reviews", []))
        all_texts = list(c.get("review_texts", []))
        # If we have fewer reviews than samples, just use what we have
        if len(reps) < SAMPLES_PER_CLUSTER and all_texts:
            extras = [t for t in all_texts if t not in reps]
            rng.shuffle(extras)
            need = SAMPLES_PER_CLUSTER - len(reps)
            reps.extend(extras[:need])
        samples = reps[:SAMPLES_PER_CLUSTER]
        while len(samples) < SAMPLES_PER_CLUSTER:
            samples.append("")

        top_asp_str = ", ".join(
            f"{a['aspect']}({a['in_cluster_pct']}%)"
            for a in c.get("top_aspects", [])
        )

        row = [
            c["cluster_id"],
            c["issue_type"],
            c.get("auto_name", ""),
            c["review_count"],
            top_asp_str,
        ] + samples + ["", ""]

        for col_i, val in enumerate(row, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=type_color.get(c["issue_type"], "FFFFFF"))
            if col_i in (1, 2, 4):  # cluster_id, issue_type, count
                cell.font = Font(bold=True)

    # Column widths
    widths = {
        "cluster_id": 11, "issue_type": 14, "auto_name": 38, "review_count": 8,
        "top_aspects": 30,
        **{f"sample_{i+1}": 50 for i in range(SAMPLES_PER_CLUSTER)},
        "coherent_yn": 12, "notes": 30,
    }
    for col_i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = widths.get(h, 15)

    # Freeze first row + first 3 cols
    ws.freeze_panes = "D2"

    # Save
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\nWrote {OUT_PATH} ({OUT_PATH.stat().st_size/1024:.1f} KB)")
    print(f"\nOpen this file in Numbers or Excel and fill in the 'coherent_yn' column.")
    print(f"After you're done, run scripts/score_cluster_validation.py to compute the purity rate.")


if __name__ == "__main__":
    main()
