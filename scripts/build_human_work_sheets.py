"""
Build all three optional-human-work spreadsheets in one go:
  1. cluster_curation.xlsx     — 100 clusters with verdict column (#11)
  2. reference_specs.xlsx       — 20 IssueSpec templates with empty fields (#10)
  3. response_ratings.xlsx      — 400 blinded (review, response) pairs (#9)

Outputs land in `human_work/` at the repo root.
"""

import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path("human_work")
OUT_DIR.mkdir(parents=True, exist_ok=True)

LABELS = ["bug_report", "feature_request", "performance", "usability",
          "compatibility", "praise", "other"]

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def header_row(ws, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER


def add_instructions_sheet(wb, title, lines):
    inst = wb.create_sheet(title, 0)
    for i, line in enumerate(lines, 1):
        inst.cell(row=i, column=1, value=line)
        if i == 1:
            inst.cell(row=i, column=1).font = Font(bold=True, size=14)
    inst.column_dimensions["A"].width = 100


# =============== #11 — Cluster Curation =================================

def build_cluster_curation():
    print("Building #11 cluster_curation.xlsx ...")
    # Use clusters_full.json (has review_texts) and merge in names/aspects
    clusters = json.load(open("data/processed/clusters_umap/clusters_full.json"))
    named = {c["cluster_id"]: c for c in
             json.load(open("data/processed/clusters_umap/clusters_named.json"))}
    for c in clusters:
        c["auto_name"] = named.get(c["cluster_id"], {}).get("auto_name", c.get("issue_type", ""))
        c["top_aspects"] = named.get(c["cluster_id"], {}).get("top_aspects", [])

    rng = random.Random(42)
    # Stratified 100 of 194 by issue_type (proportional)
    by_type = {}
    for c in clusters:
        by_type.setdefault(c["issue_type"], []).append(c)
    target_per_type = {
        "bug_report": 30, "feature_request": 30, "performance": 20,
        "usability": 16, "compatibility": 4,
    }
    selected = []
    for t, n in target_per_type.items():
        pool = sorted(by_type.get(t, []), key=lambda c: -c["review_count"])[:n]
        selected.extend(pool)

    wb = Workbook()
    wb.remove(wb.active)
    add_instructions_sheet(wb, "Instructions", [
        "CLUSTER CURATION (Task #11)",
        "",
        f"For each of {len(selected)} clusters, read the 5 sample reviews and the auto_name,",
        "then mark a verdict in the 'verdict' column:",
        "",
        "  Keep                — cluster looks coherent, name is reasonable",
        "  Split               — multiple distinct themes mixed together",
        "  Merge:c_XXXXX       — duplicate/near-dup of another cluster (give the cluster_id)",
        "  Rename:<new name>   — cluster is coherent but auto_name is wrong",
        "",
        "Optional: write a one-line reason in 'notes'.",
        "",
        "Estimated time: ~30 seconds per cluster, ~1 hour total.",
        "",
        "Output: when done, save and Anonymous will run scripts/score_cluster_curation.py to",
        "apply the verdicts and recompute cluster purity."
    ])

    ws = wb.create_sheet("Curation")
    headers = ["cluster_id", "issue_type", "auto_name", "count",
               "top_aspects", "sample_1", "sample_2", "sample_3", "sample_4", "sample_5",
               "verdict", "notes"]
    header_row(ws, headers)

    type_color = {
        "bug_report": "FCE4E4", "feature_request": "E1F0FF",
        "performance": "FFF2CC", "usability": "E2EFDA", "compatibility": "EAD8F1",
    }

    for ri, c in enumerate(selected, start=2):
        reps = c.get("representative_reviews", [])
        all_t = c.get("review_texts", reps)
        rng.shuffle(all_t)
        samples = (reps + [t for t in all_t if t not in reps])[:5]
        while len(samples) < 5:
            samples.append("")
        top_asp = ", ".join(f"{a['aspect']}({a['in_cluster_pct']}%)"
                             for a in c.get("top_aspects", []))
        row = [c["cluster_id"], c["issue_type"], c.get("auto_name", ""),
               c["review_count"], top_asp] + samples + ["", ""]
        for ci, val in enumerate(row, 1):
            cc = ws.cell(row=ri, column=ci, value=val)
            cc.alignment = Alignment(wrap_text=True, vertical="top")
            cc.border = BORDER
            cc.fill = PatternFill("solid", fgColor=type_color.get(c["issue_type"], "FFFFFF"))

    widths = {"cluster_id": 11, "issue_type": 14, "auto_name": 38, "count": 8,
              "top_aspects": 28,
              "sample_1": 50, "sample_2": 50, "sample_3": 50, "sample_4": 50, "sample_5": 50,
              "verdict": 16, "notes": 30}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 15)
    ws.freeze_panes = "D2"

    out = OUT_DIR / "cluster_curation.xlsx"
    wb.save(out)
    print(f"  saved {out} ({len(selected)} clusters)")


# =============== #10 — Reference IssueSpecs =============================

def build_reference_specs():
    print("\nBuilding #10 reference_specs.xlsx ...")
    # Use clusters_full.json (has review_texts) NOT clusters_named.json (only top-3 reps)
    clusters = json.load(open("data/processed/clusters_umap/clusters_full.json"))
    # Merge in auto_name + top_aspects from clusters_named.json
    named = {c["cluster_id"]: c for c in
             json.load(open("data/processed/clusters_umap/clusters_named.json"))}
    for c in clusters:
        c["auto_name"] = named.get(c["cluster_id"], {}).get("auto_name", c.get("issue_type", ""))
        c["top_aspects"] = named.get(c["cluster_id"], {}).get("top_aspects", [])
    auto_specs = json.load(open("data/processed/issue_specs/specs_with_taxonomy.json"))
    spec_by_cluster = {s["cluster_id"]: s for s in auto_specs}

    # Pick 20 — top by size, balanced across issue types
    target = {"bug_report": 6, "feature_request": 6, "performance": 4,
              "usability": 3, "compatibility": 1}
    by_type = {}
    for c in clusters:
        by_type.setdefault(c["issue_type"], []).append(c)
    selected = []
    for t, n in target.items():
        pool = sorted(by_type.get(t, []), key=lambda c: -c["review_count"])[:n]
        selected.extend(pool)

    wb = Workbook()
    wb.remove(wb.active)
    add_instructions_sheet(wb, "Instructions", [
        "REFERENCE ISSUE SPECS (Task #10)",
        "",
        f"For each of {len(selected)} clusters, read the 5 sample reviews and write a",
        "structured IssueSpec by hand (or edit the AI-suggested baseline shown in the",
        "'auto_baseline_*' columns).",
        "",
        "These will become the 'human-written' condition (d) in Experiment 1, the gold",
        "standard against which the LLM conditions are measured.",
        "",
        "Required fields for ALL specs:",
        "  title, description, severity (P0/P1/P2/P3), affected_component",
        "",
        "Type-specific fields (only fill the ones for the cluster's issue_type):",
        "  bug_report      → steps_to_reproduce, expected_behavior, actual_behavior",
        "  feature_request → user_story, acceptance_criteria",
        "  performance     → nfr_category (speed|battery|memory|responsiveness|scalability)",
        "  usability       → nielsen_heuristic (visibility | match_real_world |",
        "                                       user_control | consistency | error_prevention |",
        "                                       recognition_over_recall | flexibility |",
        "                                       aesthetic | error_recovery | help_documentation)",
        "  compatibility   → device_os_matrix (e.g. 'Samsung Galaxy S22, Android 13')",
        "",
        "Tips:",
        "  - Read the 5 sample reviews to understand the cluster theme.",
        "  - The 'auto_baseline_*' columns show what the AI generated — feel free to edit",
        "    or replace, but the goal is YOUR human judgment.",
        "  - For multi-line fields like steps_to_reproduce or acceptance_criteria, separate",
        "    items with ' | ' (pipe character).",
        "",
        f"Estimated time: ~5 min per cluster × {len(selected)} = ~1.5–2 hours total.",
    ])

    ws = wb.create_sheet("Specs")
    headers = ["cluster_id", "issue_type", "auto_name", "count",
               "sample_1", "sample_2", "sample_3", "sample_4", "sample_5",
               "auto_baseline_title", "auto_baseline_description",
               "your_title", "your_description", "your_severity",
               "your_affected_component",
               "your_steps_to_reproduce", "your_expected_behavior", "your_actual_behavior",
               "your_user_story", "your_acceptance_criteria",
               "your_nfr_category", "your_nielsen_heuristic", "your_device_os_matrix",
               "notes"]
    header_row(ws, headers)

    rng_extra = random.Random(99)
    for ri, c in enumerate(selected, start=2):
        reps = list(c.get("representative_reviews", []))[:3]
        # Top up to 5 by sampling from review_texts (skip duplicates)
        all_texts = list(c.get("review_texts", []))
        rng_extra.shuffle(all_texts)
        for t in all_texts:
            if len(reps) >= 5:
                break
            if t and t not in reps:
                reps.append(t)
        while len(reps) < 5:
            reps.append("")
        s = spec_by_cluster.get(c["cluster_id"], {})
        row = [c["cluster_id"], c["issue_type"], c.get("auto_name", ""),
               c["review_count"]] + reps + [
            s.get("title", ""), s.get("description", ""),
            "", "", "", "",  # title, description, severity, affected_component (you fill)
            "", "", "",      # bug fields
            "", "",          # feature fields
            "", "", "",      # performance, usability, compatibility
            "",
        ]
        for ci, val in enumerate(row, 1):
            cc = ws.cell(row=ri, column=ci, value=val)
            cc.alignment = Alignment(wrap_text=True, vertical="top")
            cc.border = BORDER

    widths = {"cluster_id": 11, "issue_type": 14, "auto_name": 30, "count": 8}
    for h in ["sample_1", "sample_2", "sample_3", "sample_4", "sample_5"]:
        widths[h] = 45
    for h in ["auto_baseline_title", "your_title"]:
        widths[h] = 35
    for h in ["auto_baseline_description", "your_description"]:
        widths[h] = 50
    widths["your_severity"] = 10
    widths["your_affected_component"] = 22
    for h in ["your_steps_to_reproduce", "your_user_story", "your_acceptance_criteria"]:
        widths[h] = 40
    for h in ["your_expected_behavior", "your_actual_behavior", "your_device_os_matrix"]:
        widths[h] = 30
    for h in ["your_nfr_category", "your_nielsen_heuristic"]:
        widths[h] = 22
    widths["notes"] = 25

    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 15)
    ws.freeze_panes = "B2"

    out = OUT_DIR / "reference_specs.xlsx"
    wb.save(out)
    print(f"  saved {out} ({len(selected)} clusters)")


# =============== #9 — Response Ratings (blinded) =========================

def build_response_ratings():
    print("\nBuilding #9 response_ratings.xlsx ...")
    paths = {
        "rrgen_baseline":      "data/processed/responses/responses_rrgen_baseline.json",
        "prompt_baseline":       "data/processed/responses/responses_prompt_baseline.json",
        "reviewagent_no_spec": "data/processed/responses/responses_reviewagent_no_spec.json",
        "reviewagent_full":    "data/processed/responses/responses_reviewagent_full.json",
    }
    cond_data = {k: json.load(open(v)) for k, v in paths.items()}

    # Build 400 rows: for each review, 4 responses BLINDED as A/B/C/D
    # Random per-review assignment so the user can't guess
    rng = random.Random(42)
    rows = []
    blinding_log = []  # cluster_id -> {A: cond, B: cond, C: cond, D: cond}

    for i in range(100):
        # one row per (review, condition); blinding shuffled per review
        condition_order = list(paths.keys())
        rng.shuffle(condition_order)
        blinding_log.append({
            "review_index": i,
            "cluster_id": cond_data["rrgen_baseline"][i]["cluster_id"],
            "blinding": {letter: cond for letter, cond in zip("ABCD", condition_order)},
        })
        for letter, cond in zip("ABCD", condition_order):
            r = cond_data[cond][i]
            rows.append({
                "review_index": i,
                "cluster_id": r["cluster_id"],
                "issue_type": r["issue_type"],
                "review_text": r["review_text"],
                "blind_id": letter,
                "response_text": r["response_text"],
                "true_condition": cond,  # not shown to user; written to blinding log
            })

    # Save blinding log (don't include in spreadsheet)
    with open(OUT_DIR / "response_ratings_blinding.json", "w") as f:
        json.dump(blinding_log, f, indent=2)

    wb = Workbook()
    wb.remove(wb.active)
    add_instructions_sheet(wb, "Instructions", [
        "RESPONSE RATINGS (Task #9)",
        "",
        "You'll see 400 rows: 100 reviews × 4 candidate responses each.",
        "The 4 candidate responses for each review are randomly labeled A/B/C/D so you",
        "can rate them without bias toward any condition.",
        "",
        "For each row, score:",
        "",
        "  quality (1-5):",
        "    5 = excellent dev-rel response, on-point and actionable",
        "    4 = good response, addresses the review well",
        "    3 = acceptable, generic but not wrong",
        "    2 = weak — too generic or partially off-target",
        "    1 = bad — irrelevant, hostile, or wrong",
        "",
        "  specificity (1-5):",
        "    5 = names the exact feature/component the user mentioned",
        "    4 = mentions a specific aspect of the issue",
        "    3 = somewhat specific but generic in places",
        "    2 = mostly generic",
        "    1 = entirely generic / could apply to any review",
        "",
        "  helpful (Y/N):",
        "    Y = the user would find this response useful in resolving their issue",
        "    N = the response would not actually help the user",
        "",
        "  notes (optional): short note for surprising cases (e.g. 'great' / 'wrong product').",
        "",
        f"Total rows: {len(rows)}",
        "Estimated time: ~30 sec per row, ~3-4 hours total. Take breaks; the cells save.",
        "",
        "When done, save the file. Anonymous will run scripts/score_response_ratings.py to",
        "unblind A/B/C/D and compute paired stats per condition.",
    ])

    ws = wb.create_sheet("Ratings")
    headers = ["review_index", "cluster_id", "issue_type", "review_text",
               "blind_id", "response_text",
               "quality_1_to_5", "specificity_1_to_5", "helpful_y_n", "notes"]
    header_row(ws, headers)

    type_color = {
        "bug_report": "FCE4E4", "feature_request": "E1F0FF",
        "performance": "FFF2CC", "usability": "E2EFDA", "compatibility": "EAD8F1",
        "praise": "FFE0CC", "other": "F0F0F0",
    }

    for ri, r in enumerate(rows, start=2):
        row = [r["review_index"], r["cluster_id"], r["issue_type"], r["review_text"],
               r["blind_id"], r["response_text"], "", "", "", ""]
        for ci, val in enumerate(row, 1):
            cc = ws.cell(row=ri, column=ci, value=val)
            cc.alignment = Alignment(wrap_text=True, vertical="top")
            cc.border = BORDER
            cc.fill = PatternFill("solid", fgColor=type_color.get(r["issue_type"], "FFFFFF"))

    widths = {"review_index": 8, "cluster_id": 11, "issue_type": 14,
              "review_text": 50, "blind_id": 9, "response_text": 60,
              "quality_1_to_5": 12, "specificity_1_to_5": 13, "helpful_y_n": 11, "notes": 25}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 15)
    ws.freeze_panes = "G2"

    out = OUT_DIR / "response_ratings.xlsx"
    wb.save(out)
    print(f"  saved {out} ({len(rows)} rows)")
    print(f"  saved {OUT_DIR / 'response_ratings_blinding.json'} (private — A/B/C/D mapping)")


if __name__ == "__main__":
    build_cluster_curation()
    build_reference_specs()
    build_response_ratings()
    print(f"\nAll three sheets in {OUT_DIR}/")
